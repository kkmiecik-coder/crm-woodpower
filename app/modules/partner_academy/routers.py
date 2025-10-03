# app/modules/partner_academy/routers.py
"""
Partner Academy Routes
======================

API endpoints i routes dla modułu PartnerAcademy.

Element 1: Recruitment
- GET /partner-academy/ - strona rekrutacyjna
- POST /partner-academy/api/application/validate - walidacja pola
- POST /partner-academy/api/application/submit - wysłanie aplikacji
- POST /partner-academy/api/application/generate-nda - generowanie PDF
- POST /partner-academy/api/application/check-email - sprawdzenie email

Element 2: Learning Platform
- GET /partner-academy/learning - platforma e-learningowa
- POST /partner-academy/api/session/init - inicjalizacja sesji
- POST /partner-academy/api/progress/load - ładowanie progressu
- POST /partner-academy/api/progress/update - aktualizacja progressu
- POST /partner-academy/api/time/sync - synchronizacja czasu
- POST /partner-academy/api/quiz/validate - walidacja quizu

Admin Panel:
- GET /partner-academy/admin/ - panel administracyjny
- GET /partner-academy/admin/api/stats - statystyki
- GET /partner-academy/admin/api/applications - lista aplikacji z filtrowaniem
- GET /partner-academy/admin/api/application/<id> - szczegóły aplikacji
- POST /partner-academy/admin/api/application/<id>/status - zmiana statusu
- POST /partner-academy/admin/api/application/<id>/note - dodanie notatki
- GET /partner-academy/admin/api/learning-sessions/<session_id> - szczegóły sesji
- GET /partner-academy/admin/api/export - eksport do XLSX

Autor: Development Team
Data: 2025-09-30
"""

from flask import render_template, request, jsonify, current_app, send_file, session, redirect, url_for, flash, make_response
from modules.partner_academy import partner_academy_bp
from modules.partner_academy.services import ApplicationService, EmailService, LearningService
from modules.partner_academy.validators import validate_application_data, validate_file, validate_quiz_answers
from modules.partner_academy.utils import rate_limit, get_quiz_answers, generate_nda_pdf
from extensions import db
from modules.partner_academy.models import PartnerApplication, PartnerLearningSession
import io
import os
from datetime import datetime, timedelta

from sqlalchemy import func, or_, desc
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from functools import wraps

def json_response(data, status=200):
    """Helper do tworzenia JSON response z właściwym Content-Type"""
    response = make_response(jsonify(data), status)
    response.headers['Content-Type'] = 'application/json; charset=utf-8'
    return response

def login_required(func):
    """Dekorator wymagający zalogowania dla panelu admina"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        user_email = session.get('user_email')
        if not user_email:
            flash("Twoja sesja wygasła. Zaloguj się ponownie.", "info")
            return redirect(url_for('login'))
        return func(*args, **kwargs)
    return wrapper


# ============================================================================
# ELEMENT 1: RECRUITMENT - VIEWS
# ============================================================================

@partner_academy_bp.route('/')
def recruitment():
    """Strona rekrutacyjna (Element 1)"""
    return render_template('recruitment.html')


# ============================================================================
# ELEMENT 2: LEARNING PLATFORM - VIEWS
# ============================================================================

@partner_academy_bp.route('/learning')
def learning():
    """Platforma e-learningowa (Element 2)"""
    return render_template('learning_platform.html')


# ============================================================================
# API ENDPOINTS - ELEMENT 1: RECRUITMENT
# ============================================================================

@partner_academy_bp.route('/api/application/validate', methods=['POST'])
def validate_application_field():
    """
    Walidacja pojedynczego pola formularza (AJAX)
    
    Request JSON:
    {
        "field_name": "email",
        "field_value": "test@example.com"
    }
    
    Response:
    {
        "success": true,
        "valid": true,
        "error": null
    }
    """
    try:
        data = request.get_json()
        field_name = data.get('field_name')
        field_value = data.get('field_value')
        
        if not field_name:
            return jsonify({
                'success': False,
                'error': 'Brak nazwy pola'
            }), 400
        
        # Waliduj pojedyncze pole
        temp_form_data = {field_name: field_value}
        is_valid, errors = validate_application_data(temp_form_data)
        
        field_error = errors.get(field_name)
        
        return jsonify({
            'success': True,
            'valid': field_error is None,
            'error': field_error
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Validation error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Błąd walidacji'
        }), 500


@partner_academy_bp.route('/api/application/submit', methods=['POST'])
def submit_application():
    """
    Wysłanie kompletnego formularza aplikacyjnego z plikiem NDA
    
    Request (multipart/form-data):
    - Wszystkie pola formularza
    - nda_file: plik NDA
    
    Response:
    {
        "success": true,
        "message": "Aplikacja została wysłana",
        "application_id": 123
    }
    """
    try:
        # Pobierz dane z formularza
        form_data = request.form.to_dict()
        
        # Pobierz plik NDA
        nda_file = request.files.get('nda_file')
        
        current_app.logger.info(f"Received application from: {form_data.get('email')}")
        
        # ========================================================================
        # WALIDACJA DANYCH FORMULARZA
        # ========================================================================
        
        is_valid, errors = validate_application_data(form_data)

        if not is_valid:
            current_app.logger.warning(f"Validation errors: {errors}")
            return json_response({
                'success': False,
                'errors': errors
            }, 400)
        
        # ========================================================================
        # WALIDACJA PLIKU NDA
        # ========================================================================
        
        file_valid, file_error = validate_file(nda_file)
        
        if not file_valid:
            current_app.logger.warning(f"File validation error: {file_error}")
            return json_response({
                'success': False,
                'error': file_error
            }, 400)
        
        # ========================================================================
        # SPRAWDŹ CZY EMAIL JUŻ ISTNIEJE
        # ========================================================================
        
        existing_application = ApplicationService.get_application_by_email(
            form_data['email']
        )
        
        if existing_application:
            current_app.logger.warning(
                f"Duplicate application attempt: {form_data['email']}"
            )
            return json_response({
                'success': False,
                'error': 'Aplikacja z tym adresem email już istnieje'
            }, 400)
        
        # ========================================================================
        # UTWÓRZ APLIKACJĘ
        # ========================================================================
        
        # Pobierz IP i user agent
        ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
        if ip_address and ',' in ip_address:
            ip_address = ip_address.split(',')[0].strip()
        
        user_agent = request.headers.get('User-Agent', '')
        
        # Utwórz aplikację w bazie
        application = ApplicationService.create_application(
            form_data=form_data,
            file=nda_file,
            ip_address=ip_address,
            user_agent=user_agent
        )
        
        current_app.logger.info(
            f"Application created successfully: ID={application.id}, Email={application.email}"
        )
        
        # ========================================================================
        # WYSYŁKA EMAILI
        # ========================================================================
        
        try:
            # Email do kandydata
            EmailService.send_application_confirmation(application)
            
            # Email do admina
            EmailService.send_admin_notification(application)
            
        except Exception as email_error:
            current_app.logger.error(f"Email sending failed: {str(email_error)}")
            # Nie przerywamy procesu jeśli email się nie wyśle
        
        # ========================================================================
        # RESPONSE
        # ========================================================================
        
        return json_response({
            'success': True,
            'message': 'Aplikacja została wysłana pomyślnie',
            'application_id': application.id
        }, 201)
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Application submission error: {str(e)}", exc_info=True)
        
        return json_response({
            'success': False,
            'error': 'Wystąpił błąd podczas przetwarzania aplikacji. Spróbuj ponownie.'
        }, 500)


@partner_academy_bp.route('/api/application/generate-nda', methods=['POST'])
def generate_nda():
    """
    Generowanie PDF z umową NDA na podstawie danych z formularza
    
    Request JSON:
    {
        "first_name": "Jan",
        "last_name": "Kowalski",
        "email": "jan@example.com",
        "city": "Warszawa",
        ... (wszystkie dane formularza)
    }
    
    Response:
    - PDF file (application/pdf)
    """
    try:
        data = request.get_json()
        
        # Waliduj podstawowe dane
        required_fields = ['first_name', 'last_name', 'email', 'city', 'pesel']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return jsonify({
                'success': False,
                'error': f'Brak wymaganych pól: {", ".join(missing_fields)}'
            }), 400
        
        current_app.logger.info(f"Generating NDA for: {data.get('email')}")
        
        # Generuj PDF - zwróci surowe bajty
        pdf_bytes = generate_nda_pdf(data)
        
        if not pdf_bytes:
            return jsonify({
                'success': False,
                'error': 'Nie udało się wygenerować PDF'
            }), 500
        
        # Przygotuj nazwę pliku
        filename = f"NDA_{data['last_name']}_{data['first_name']}.pdf"
        
        current_app.logger.info(f"NDA generated successfully: {filename}")
        
        # Zwróć PDF używając Response z surowymi bajtami
        from flask import Response
        return Response(
            pdf_bytes,
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )
        
    except Exception as e:
        current_app.logger.error(f"NDA generation error: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Wystąpił błąd podczas generowania PDF'
        }), 500


@partner_academy_bp.route('/api/application/check-email', methods=['POST'])
def check_email_exists():
    """
    Sprawdź czy email już istnieje w bazie (pomocnicze API)
    
    Request JSON:
    {
        "email": "test@example.com"
    }
    
    Response:
    {
        "success": true,
        "exists": false
    }
    """
    try:
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({
                'success': False,
                'error': 'Brak adresu email'
            }), 400
        
        existing = ApplicationService.get_application_by_email(email)
        
        return jsonify({
            'success': True,
            'exists': existing is not None
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Email check error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Błąd sprawdzania email'
        }), 500


# ============================================================================
# API ENDPOINTS - ELEMENT 2: LEARNING PLATFORM
# ============================================================================

@partner_academy_bp.route('/api/session/init', methods=['POST'])
def init_session():
    """Inicjalizacja sesji - znajdź istniejącą po IP lub utwórz nową"""
    try:
        # Pobierz IP użytkownika
        ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
        if ip_address and ',' in ip_address:
            ip_address = ip_address.split(',')[0].strip()
        
        # Znajdź lub utwórz sesję
        session_id = LearningService.find_or_create_session_by_ip(ip_address)
        
        return jsonify({
            'success': True,
            'session_id': session_id,
            'ip_address': ip_address
        }), 200
    
    except Exception as e:
        current_app.logger.error(f"Session init error: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Błąd inicjalizacji sesji'
        }), 500


@partner_academy_bp.route('/api/progress/load', methods=['POST'])
def load_progress():
    """Ładowanie zapisanego progressu użytkownika"""
    try:
        data = request.get_json()
        session_id = data.get('session_id')
        
        if not session_id:
            return jsonify({
                'success': False,
                'message': 'Brak session_id'
            }), 400
        
        progress = LearningService.load_progress(session_id)
        
        return jsonify({
            'success': True,
            'progress': progress
        }), 200
    
    except Exception as e:
        current_app.logger.error(f"Load progress error: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Błąd ładowania progressu'
        }), 500


@partner_academy_bp.route('/api/progress/update', methods=['POST'])
def update_progress():
    """Aktualizacja progressu użytkownika"""
    try:
        data = request.get_json()
        session_id = data.get('session_id')
        current_step = data.get('current_step')
        completed_steps = data.get('completed_steps', [])
        quiz_results = data.get('quiz_results')
        
        if not session_id or not current_step:
            return jsonify({
                'success': False,
                'message': 'Brak wymaganych danych'
            }), 400
        
        LearningService.update_progress(
            session_id=session_id,
            current_step=current_step,
            completed_steps=completed_steps,
            quiz_results=quiz_results
        )
        
        return jsonify({
            'success': True,
            'message': 'Progress zapisany'
        }), 200
    
    except Exception as e:
        current_app.logger.error(f"Update progress error: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Błąd aktualizacji progressu'
        }), 500


@partner_academy_bp.route('/api/time/sync', methods=['POST'])
def sync_time():
    """Synchronizacja czasu spędzonego w aplikacji"""
    try:
        data = request.get_json()
        session_id = data.get('session_id')
        time_spent = data.get('time_spent')
        step_time_tracking = data.get('step_time_tracking')
        
        if not session_id or time_spent is None:
            return jsonify({
                'success': False,
                'message': 'Brak wymaganych danych'
            }), 400
        
        LearningService.sync_time(
            session_id=session_id,
            time_spent=time_spent,
            step_time_tracking=step_time_tracking
        )
        
        return jsonify({
            'success': True,
            'message': 'Czas zsynchronizowany'
        }), 200
    
    except Exception as e:
        current_app.logger.error(f"Sync time error: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Błąd synchronizacji czasu'
        }), 500


@partner_academy_bp.route('/api/quiz/validate', methods=['POST'])
def validate_quiz():
    """Walidacja odpowiedzi w quizie"""
    try:
        data = request.get_json()
        step = data.get('step')
        answers = data.get('answers')
        
        if not step or not answers:
            return jsonify({
                'success': False,
                'message': 'Brak wymaganych danych'
            }), 400
        
        validation = validate_quiz_answers(step, answers)
        
        return jsonify({
            'success': True,
            'all_correct': validation['all_correct'],
            'results': validation['results']
        }), 200
    
    except Exception as e:
        current_app.logger.error(f"Validate quiz error: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Błąd walidacji quizu'
        }), 500


# ============================================================================
# ADMIN PANEL - VIEWS
# ============================================================================

@partner_academy_bp.route('/admin/')
@login_required
def admin_dashboard():
    """Strona główna panelu admina"""
    return render_template('admin.html')


# ============================================================================
# ADMIN PANEL - API ENDPOINTS
# ============================================================================

@partner_academy_bp.route('/admin/api/stats')
@login_required
def get_admin_stats():
    """Statystyki aplikacji i szkoleń dla dashboardu"""
    try:
        # ============================================================================
        # STATYSTYKI APLIKACJI REKRUTACYJNYCH
        # ============================================================================
        total_applications = PartnerApplication.query.count()
        pending_count = PartnerApplication.query.filter_by(status='pending').count()
        contacted_count = PartnerApplication.query.filter_by(status='contacted').count()
        accepted_count = PartnerApplication.query.filter_by(status='accepted').count()
        rejected_count = PartnerApplication.query.filter_by(status='rejected').count()
        
        # ============================================================================
        # STATYSTYKI SESJI SZKOLENIOWYCH
        # ============================================================================
        total_sessions = PartnerLearningSession.query.count()
        
        # Sesje ukończone (completed_at IS NOT NULL)
        completed_sessions = PartnerLearningSession.query.filter(
            PartnerLearningSession.completed_at.isnot(None)
        ).count()
        
        # Sesje aktywne (ostatnia aktywność w ciągu ostatnich 7 dni)
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        active_sessions = PartnerLearningSession.query.filter(
            PartnerLearningSession.last_activity_at >= seven_days_ago,
            PartnerLearningSession.completed_at.is_(None)
        ).count()
        
        # Średni postęp (% ukończonych kroków)
        all_sessions = PartnerLearningSession.query.all()
        if all_sessions:
            total_progress = 0
            for session in all_sessions:
                completed_steps = session.completed_steps or []
                progress_percent = (len(completed_steps) / 29) * 100
                total_progress += progress_percent
            avg_progress = round(total_progress / len(all_sessions), 1)
        else:
            avg_progress = 0
        
        return jsonify({
            'success': True,
            'data': {
                # Statystyki aplikacji
                'total_applications': total_applications,
                'pending_count': pending_count,
                'contacted_count': contacted_count,
                'accepted_count': accepted_count,
                'rejected_count': rejected_count,
                
                # Statystyki szkoleń
                'total_sessions': total_sessions,
                'active_sessions': active_sessions,
                'completed_sessions': completed_sessions,
                'avg_progress': avg_progress
            }
        }), 200
    except Exception as e:
        current_app.logger.error(f"Admin stats error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd pobierania statystyk'}), 500


@partner_academy_bp.route('/admin/api/applications')
@login_required
def get_admin_applications():
    """Lista aplikacji z filtrowaniem i paginacją"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        status_filter = request.args.get('status', '')
        search = request.args.get('search', '')
        is_b2b = request.args.get('is_b2b', '')
        
        query = PartnerApplication.query
        
        # Filtrowanie po statusie
        if status_filter:
            query = query.filter_by(status=status_filter)
        
        # Filtrowanie B2B/B2C
        if is_b2b == 'true':
            query = query.filter_by(is_b2b=True)
        elif is_b2b == 'false':
            query = query.filter_by(is_b2b=False)
        
        # Wyszukiwanie
        if search:
            pattern = f'%{search}%'
            query = query.filter(
                or_(
                    PartnerApplication.first_name.ilike(pattern),
                    PartnerApplication.last_name.ilike(pattern),
                    PartnerApplication.email.ilike(pattern),
                    PartnerApplication.phone.ilike(pattern),
                    PartnerApplication.company_name.ilike(pattern),
                    PartnerApplication.nip.ilike(pattern)
                )
            )
        
        # Paginacja
        pagination = query.order_by(desc(PartnerApplication.created_at)).paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        
        # POPRAWIONY RETURN - NOWA STRUKTURA
        applications = []
        for app in pagination.items:
            app_data = {
                'id': app.id,
                'first_name': app.first_name,
                'last_name': app.last_name,
                'email': app.email,
                'phone': app.phone,
                'city': app.city,
                'address': app.address,
                'postal_code': app.postal_code,
                'status': app.status,
                'is_b2b': app.is_b2b,
                'created_at': app.created_at.strftime('%Y-%m-%d %H:%M') if app.created_at else None,
                'has_nda_file': bool(app.nda_filepath)
            }
            
            # Dodaj dane B2B jeśli istnieją
            if app.is_b2b:
                app_data['company_name'] = app.company_name
                app_data['nip'] = app.nip
            
            applications.append(app_data)
        
        return jsonify({
            'success': True,
            'data': {
                'applications': applications,
                'pagination': {
                    'page': pagination.page,
                    'per_page': pagination.per_page,
                    'total': pagination.total,
                    'pages': pagination.pages
                }
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Admin applications list error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd pobierania listy aplikacji'}), 500


@partner_academy_bp.route('/admin/api/application/<int:application_id>')
@login_required
def get_admin_application_detail(application_id):
    """Szczegóły pojedynczej aplikacji (BEZ powiązania z sesją szkoleniową)"""
    try:
        app = PartnerApplication.query.get_or_404(application_id)
        
        # Podstawowe dane
        detail = {
            'id': app.id,
            'first_name': app.first_name,
            'last_name': app.last_name,
            'email': app.email,
            'phone': app.phone,
            'city': app.city,
            'address': app.address,
            'postal_code': app.postal_code,
            'pesel': app.pesel,
            'voivodeship': app.voivodeship,
            'business_location': app.business_location,
            'about_text': app.about_text,
            'status': app.status,
            'is_b2b': app.is_b2b,
            'data_processing_consent': app.data_processing_consent,
            'created_at': app.created_at.strftime('%Y-%m-%d %H:%M:%S') if app.created_at else None,
            'updated_at': app.updated_at.strftime('%Y-%m-%d %H:%M:%S') if app.updated_at else None,
            'ip_address': app.ip_address,
            'user_agent': app.user_agent,
            'notes': json.loads(app.notes) if app.notes else [],
            'has_nda_file': bool(app.nda_filepath)
        }
        
        # Dane NDA
        if app.nda_filepath:
            detail['nda_filename'] = app.nda_filename
            detail['nda_filesize'] = app.nda_filesize
            detail['nda_mime_type'] = app.nda_mime_type
        
        # Dane B2B
        if app.is_b2b:
            detail['company_name'] = app.company_name
            detail['nip'] = app.nip
            detail['regon'] = app.regon
            detail['company_address'] = app.company_address
            detail['company_city'] = app.company_city
            detail['company_postal_code'] = app.company_postal_code
        
        # NIE ŁĄCZYMY Z SESJAMI - PIN jest wspólny dla wszystkich
        detail['learning_session'] = None
        
        return jsonify({
            'success': True,
            'data': detail
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Admin application detail error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd pobierania szczegółów aplikacji'}), 500


@partner_academy_bp.route('/admin/api/learning-sessions')
@login_required
def get_admin_learning_sessions():
    """Lista sesji szkoleniowych z filtrowaniem i paginacją"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        status = request.args.get('status', '')
        search = request.args.get('search', '')
        
        query = PartnerLearningSession.query
        
        # Wyszukiwanie po session_id
        if search:
            query = query.filter(PartnerLearningSession.session_id.ilike(f'%{search}%'))
        
        # Filtrowanie po statusie
        if status == 'active':
            seven_days_ago = datetime.utcnow() - timedelta(days=7)
            query = query.filter(
                PartnerLearningSession.last_accessed_at >= seven_days_ago,
                PartnerLearningSession.completed_at.is_(None)
            )
        elif status == 'completed':
            query = query.filter(PartnerLearningSession.completed_at.isnot(None))
        elif status == 'inactive':
            seven_days_ago = datetime.utcnow() - timedelta(days=7)
            query = query.filter(
                or_(
                    PartnerLearningSession.last_accessed_at < seven_days_ago,
                    PartnerLearningSession.last_accessed_at.is_(None)
                ),
                PartnerLearningSession.completed_at.is_(None)
            )
        
        # Sortowanie (najnowsze na górze)
        query = query.order_by(desc(PartnerLearningSession.created_at))
        
        # Paginacja
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        # Formatowanie wyników
        sessions = []
        for session in pagination.items:
            completed_steps = session.completed_steps or []
            total_time_hours = round(session.total_time_spent / 3600, 2)
            
            # Sprawdź czy sesja jest aktywna
            is_active = False
            if session.last_accessed_at:
                seven_days_ago = datetime.utcnow() - timedelta(days=7)
                is_active = session.last_accessed_at >= seven_days_ago
            
            sessions.append({
                'id': session.id,
                'email': session.session_id,  # Używamy session_id jako identyfikatora
                'current_step': session.current_step,
                'completed_steps_count': len(completed_steps),
                'total_time_hours': total_time_hours,
                'last_activity_at': session.last_accessed_at.strftime('%Y-%m-%d %H:%M') if session.last_accessed_at else None,
                'is_completed': session.completed_at is not None,
                'is_active': is_active
            })
        
        return jsonify({
            'success': True,
            'data': {
                'sessions': sessions,
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': pagination.total,
                    'pages': pagination.pages
                }
            }
        }), 200
    except Exception as e:
        current_app.logger.error(f"Admin sessions error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd pobierania sesji'}), 500


@partner_academy_bp.route('/admin/api/learning-session/<int:session_id>')
@login_required
def get_admin_learning_session_detail(session_id):
    """Szczegóły pojedynczej sesji szkoleniowej"""
    try:
        session_obj = PartnerLearningSession.query.get_or_404(session_id)
        
        completed_steps = session_obj.completed_steps or []
        locked_steps = session_obj.locked_steps or []
        quiz_results = session_obj.quiz_results or {}
        step_times = session_obj.step_times or {}
        
        total_time_hours = round(session_obj.total_time_spent / 3600, 2)
        
        # Sprawdź czy sesja jest aktywna
        is_active = False
        if session_obj.last_accessed_at:
            seven_days_ago = datetime.utcnow() - timedelta(days=7)
            is_active = session_obj.last_accessed_at >= seven_days_ago
        
        # NIE ŁĄCZYMY Z APLIKACJAMI - brak wspólnego klucza
        
        return jsonify({
            'success': True,
            'data': {
                'id': session_obj.id,
                'email': session_obj.session_id,  # Używamy session_id jako identyfikatora
                'pin_code': 'N/A',  # PIN jest wspólny
                'current_step': session_obj.current_step,
                'completed_steps': completed_steps,
                'locked_steps': locked_steps,
                'quiz_results': quiz_results,
                'step_times': step_times,
                'total_time_hours': total_time_hours,
                'is_completed': session_obj.completed_at is not None,
                'is_active': is_active,
                'created_at': session_obj.created_at.strftime('%Y-%m-%d %H:%M') if session_obj.created_at else None,
                'last_activity_at': session_obj.last_accessed_at.strftime('%Y-%m-%d %H:%M') if session_obj.last_accessed_at else None,
                'completed_at': session_obj.completed_at.strftime('%Y-%m-%d %H:%M') if session_obj.completed_at else None,
                'application_id': None  # Nie łączymy
            }
        }), 200
    except Exception as e:
        current_app.logger.error(f"Admin session detail error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd pobierania sesji'}), 500

@partner_academy_bp.route('/admin/api/export-applications')
@login_required
def export_applications_xlsx():
    """Eksport aplikacji do pliku XLSX"""
    try:
        status = request.args.get('status', '')
        search = request.args.get('search', '')
        
        query = PartnerApplication.query
        
        if status:
            query = query.filter_by(status=status)
        if search:
            query = query.filter(
                or_(
                    PartnerApplication.first_name.ilike(f'%{search}%'),
                    PartnerApplication.last_name.ilike(f'%{search}%'),
                    PartnerApplication.email.ilike(f'%{search}%')
                )
            )
        
        applications = query.order_by(desc(PartnerApplication.created_at)).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Aplikacje"
        
        headers = [
            'ID', 'Data utworzenia', 'Imię', 'Nazwisko', 'Email', 'Telefon',
            'Miasto', 'Adres', 'Kod pocztowy', 'Województwo', 'Status', 'B2B', 
            'Nazwa firmy', 'NIP'
        ]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="ED6B24", end_color="ED6B24", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for app in applications:
            row = [
                app.id,
                app.created_at.strftime('%Y-%m-%d %H:%M') if app.created_at else '',
                app.first_name,
                app.last_name,
                app.email,
                app.phone,
                app.city,
                app.address,
                app.postal_code,
                app.voivodeship,
                app.status,
                'Tak' if app.is_b2b else 'Nie',
                app.company_name if app.is_b2b else '',
                app.nip if app.is_b2b else ''
            ]
            ws.append(row)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f'aplikacje_rekrutacyjne_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f"Export applications error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd eksportu'}), 500


@partner_academy_bp.route('/admin/api/export-sessions')
@login_required
def export_sessions_xlsx():
    """Eksport sesji szkoleniowych do pliku XLSX"""
    try:
        status = request.args.get('status', '')
        search = request.args.get('search', '')
        
        query = PartnerLearningSession.query
        
        if search:
            query = query.filter(PartnerLearningSession.session_id.ilike(f'%{search}%'))
        
        if status == 'active':
            seven_days_ago = datetime.utcnow() - timedelta(days=7)
            query = query.filter(
                PartnerLearningSession.last_accessed_at >= seven_days_ago,
                PartnerLearningSession.completed_at.is_(None)
            )
        elif status == 'completed':
            query = query.filter(PartnerLearningSession.completed_at.isnot(None))
        elif status == 'inactive':
            seven_days_ago = datetime.utcnow() - timedelta(days=7)
            query = query.filter(
                or_(
                    PartnerLearningSession.last_accessed_at < seven_days_ago,
                    PartnerLearningSession.last_accessed_at.is_(None)
                ),
                PartnerLearningSession.completed_at.is_(None)
            )
        
        sessions = query.order_by(desc(PartnerLearningSession.created_at)).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sesje Szkoleniowe"
        
        headers = [
            'ID', 'Session ID', 'Aktualny krok', 'Ukończone kroki',
            '% Postępu', 'Czas spędzony (h)', 'Data rozpoczęcia',
            'Ostatnia aktywność', 'Data ukończenia', 'Status'
        ]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for session in sessions:
            completed_steps = session.completed_steps or []
            progress_percent = round((len(completed_steps) / 29) * 100, 1)
            total_time_hours = round(session.total_time_spent / 3600, 2)
            
            if session.completed_at:
                status_text = 'Ukończone'
            elif session.last_accessed_at:
                seven_days_ago = datetime.utcnow() - timedelta(days=7)
                status_text = 'Aktywne' if session.last_accessed_at >= seven_days_ago else 'Nieaktywne'
            else:
                status_text = 'Nieaktywne'
            
            row = [
                session.id,
                session.session_id,
                session.current_step,
                len(completed_steps),
                f'{progress_percent}%',
                total_time_hours,
                session.created_at.strftime('%Y-%m-%d %H:%M') if session.created_at else '',
                session.last_accessed_at.strftime('%Y-%m-%d %H:%M') if session.last_accessed_at else '',
                session.completed_at.strftime('%Y-%m-%d %H:%M') if session.completed_at else '',
                status_text
            ]
            ws.append(row)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f'sesje_szkoleniowe_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f"Export sessions error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd eksportu'}), 500
@partner_academy_bp.route('/admin/api/application/<int:application_id>/status', methods=['POST'])
@login_required
def update_application_status(application_id):
    """Zmiana statusu aplikacji"""
    try:
        data = request.get_json()
        new_status = data.get('status')
        notes = data.get('notes')
        
        if not new_status:
            return jsonify({'success': False, 'message': 'Brak nowego statusu'}), 400
        
        application = ApplicationService.update_application_status(
            application_id=application_id,
            new_status=new_status,
            notes=notes
        )
        
        # Wyślij email do kandydata o zmianie statusu
        try:
            EmailService.send_status_update(application, new_status)
        except Exception as email_error:
            current_app.logger.error(f"Email sending failed: {str(email_error)}")
        
        return jsonify({
            'success': True,
            'message': 'Status zaktualizowany',
            'new_status': application.status
        }), 200
        
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Admin status update error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd aktualizacji statusu'}), 500


@partner_academy_bp.route('/admin/api/application/<int:application_id>/note', methods=['POST'])
@login_required
def add_application_note(application_id):
    """Dodanie notatki do aplikacji"""
    try:
        data = request.get_json()
        note_text = data.get('note')
        
        if not note_text:
            return jsonify({'success': False, 'message': 'Brak treści notatki'}), 400
        
        app = PartnerApplication.query.get_or_404(application_id)
        
        # Dodaj notatkę
        current_notes = json.loads(app.notes) if app.notes else []
        new_note = {
            'timestamp': datetime.utcnow().isoformat(),
            'author': session.get('user_email', 'admin'),
            'text': note_text
        }
        current_notes.append(new_note)
        
        app.notes = json.dumps(current_notes)
        app.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Notatka dodana', 'note': new_note}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Admin note add error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd dodawania notatki'}), 500


@partner_academy_bp.route('/admin/api/application/<int:application_id>/nda')
@login_required
def download_nda(application_id):
    """Pobieranie pliku NDA"""
    try:
        filepath = ApplicationService.get_nda_file_path(application_id)
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'success': False, 'message': 'Plik NDA nie istnieje'}), 404
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=os.path.basename(filepath)
        )
        
    except Exception as e:
        current_app.logger.error(f"NDA download error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd pobierania pliku'}), 500


@partner_academy_bp.route('/admin/api/learning-sessions/<session_id>')
@login_required
def get_admin_learning_session(session_id):
    """Szczegóły postępu w szkoleniu"""
    try:
        session_obj = PartnerLearningSession.query.filter_by(session_id=session_id).first_or_404()
        
        return jsonify({
            'success': True,
            'data': {
                'session_id': session_obj.session_id,
                'current_step': session_obj.current_step,
                'completed_steps': session_obj.completed_steps or [],
                'quiz_results': session_obj.quiz_results or {},
                'total_time_spent': session_obj.total_time_spent,
                'total_hours': round(session_obj.total_time_spent / 3600, 2),
                'is_completed': session_obj.is_completed,
                'completed_at': session_obj.completed_at.strftime('%Y-%m-%d %H:%M') if session_obj.completed_at else None,
                'last_accessed_at': session_obj.last_accessed_at.strftime('%Y-%m-%d %H:%M') if session_obj.last_accessed_at else None
            }
        }), 200
    except Exception as e:
        current_app.logger.error(f"Admin session error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd pobierania postępu'}), 500


@partner_academy_bp.route('/admin/api/export')
@login_required
def export_admin_applications():
    """Eksport aplikacji do XLSX z wszystkimi polami"""
    try:
        status_filter = request.args.get('status', '')
        search = request.args.get('search', '')
        is_b2b = request.args.get('is_b2b', '')
        
        query = PartnerApplication.query
        
        if status_filter:
            query = query.filter_by(status=status_filter)
        
        if is_b2b == 'true':
            query = query.filter_by(is_b2b=True)
        elif is_b2b == 'false':
            query = query.filter_by(is_b2b=False)
        
        if search:
            pattern = f'%{search}%'
            query = query.filter(
                or_(
                    PartnerApplication.first_name.ilike(pattern),
                    PartnerApplication.last_name.ilike(pattern),
                    PartnerApplication.email.ilike(pattern)
                )
            )
        
        applications = query.order_by(desc(PartnerApplication.created_at)).all()
        
        # Utwórz workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Aplikacje"
        
        # Nagłówki
        headers = [
            'ID', 'Imię', 'Nazwisko', 'Email', 'Telefon', 
            'Miasto', 'Adres', 'Kod pocztowy', 'PESEL',
            'Status', 'Typ', 'Data utworzenia',
            'Firma', 'NIP', 'REGON', 'Adres firmy', 'Miasto firmy', 'Kod pocztowy firmy',
            'O sobie', 'IP', 'Ma NDA'
        ]
        
        # Stylowanie nagłówków
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="ED6B24", end_color="ED6B24", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dane
        for row_num, app in enumerate(applications, 2):
            ws.cell(row=row_num, column=1, value=app.id)
            ws.cell(row=row_num, column=2, value=app.first_name)
            ws.cell(row=row_num, column=3, value=app.last_name)
            ws.cell(row=row_num, column=4, value=app.email)
            ws.cell(row=row_num, column=5, value=app.phone)
            ws.cell(row=row_num, column=6, value=app.city)
            ws.cell(row=row_num, column=7, value=app.address)
            ws.cell(row=row_num, column=8, value=app.postal_code)
            ws.cell(row=row_num, column=9, value=app.pesel)
            ws.cell(row=row_num, column=10, value=app.status)
            ws.cell(row=row_num, column=11, value='B2B' if app.is_b2b else 'B2C')
            ws.cell(row=row_num, column=12, value=app.created_at.strftime('%Y-%m-%d %H:%M') if app.created_at else '')
            
            # Dane B2B - wszystkie kolumny przesunięte o 1
            ws.cell(row=row_num, column=13, value=app.company_name if app.is_b2b else '')
            ws.cell(row=row_num, column=14, value=app.nip if app.is_b2b else '')
            ws.cell(row=row_num, column=15, value=app.regon if app.is_b2b else '')
            ws.cell(row=row_num, column=16, value=app.company_address if app.is_b2b else '')
            ws.cell(row=row_num, column=17, value=app.company_city if app.is_b2b else '')
            ws.cell(row=row_num, column=18, value=app.company_postal_code if app.is_b2b else '')
            
            ws.cell(row=row_num, column=19, value=app.about_text or '')
            ws.cell(row=row_num, column=20, value=app.ip_address or '')
            ws.cell(row=row_num, column=21, value='Tak' if app.nda_filepath else 'Nie')
        
        # Autosize kolumn
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Zapisz do bufora
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'aplikacje_partner_academy_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f"Export error: {str(e)}")
        return jsonify({'success': False, 'message': 'Błąd eksportu'}), 500