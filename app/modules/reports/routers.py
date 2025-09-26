# modules/reports/routers.py
"""
Endpointy Flask dla modułu Reports
"""
import csv
import re
import os
os.environ['OPENBLAS_NUM_THREADS'] = '1'
import pandas as pd
import io
import sys
from flask import render_template, jsonify, request, session, redirect, url_for, flash, Response, make_response
from datetime import datetime, timedelta, date
from functools import wraps
from extensions import db
from . import reports_bp
from .models import BaselinkerReportOrder, ReportsSyncLog
from .service import BaselinkerReportsService, get_reports_service
from modules.logging import get_structured_logger
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import Dict, Optional, Tuple, List

# Inicjalizacja loggera
reports_logger = get_structured_logger('reports.routers')
reports_logger.info("✅ reports_logger zainicjowany poprawnie w routers.py")

def login_required(func):
    """Dekorator wymagający zalogowania"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        user_email = session.get('user_email')
        if not user_email:
            reports_logger.warning("Próba dostępu bez autoryzacji",
                                 endpoint=request.endpoint,
                                 ip=request.remote_addr)
            flash("Twoja sesja wygasła. Zaloguj się ponownie.", "error")
            return redirect(url_for('login'))
        return func(*args, **kwargs)
    return wrapper

def generate_product_key_router(order_id, product, product_index=None):
    """
    ✅ POPRAWIONA FUNKCJA: Identyczna logika jak w frontendzie volume_manager.js
    Pomocnicza funkcja do generowania kluczy produktów - musi być IDENTYCZNA z frontendem
    """
    order_product_id = product.get('order_product_id')
    product_id_raw = product.get('product_id')
    
    # ✅ KLUCZOWA ZMIANA: Sprawdź czy product_id jest RZECZYWIŚCIE pusty
    # W danych widzimy product_id: '' (pusty string), nie None
    is_product_id_empty = not product_id_raw or str(product_id_raw).strip() == '' or str(product_id_raw) == 'unknown'
    
    # PRIORYTET 1: order_product_id TYLKO gdy product_id nie jest pusty
    # (to różni się od poprzedniej logiki!)
    if order_product_id and str(order_product_id).strip() and not is_product_id_empty:
        return f"{order_id}_{order_product_id}"
    
    # PRIORYTET 2: product_id (jeśli nie jest pusty)
    elif not is_product_id_empty:
        return f"{order_id}_{product_id_raw}"
    
    # ✅ PRIORYTET 3: product_index z prefiksem "idx_" (gdy product_id jest pusty)
    elif product_index is not None:
        return f"{order_id}_idx_{product_index}"
    
    # OSTATECZNOŚĆ: 'unknown' (może powodować konflikty)
    else:
        return f"{order_id}_unknown"

@reports_bp.route('/')
@login_required
def reports_home():
    """
    Główna strona modułu Reports
    """
    user_email = session.get('user_email')
    reports_logger.info("Dostęp do modułu Reports", user_email=user_email)
    
    try:
        # ZMIANA: Sprawdź wszystkie nowe zamówienia, nie tylko z ostatnich 48h
        service = get_reports_service()
        
        # Sprawdź nowe zamówienia z ostatnich 4 dni
        try:
            # Pobierz wszystkie zamówienia z Baselinker
            date_from_4_days = datetime.now() - timedelta(days=4)
            orders = service.fetch_orders_from_baselinker(date_from=date_from_4_days)
            
            if orders:
                # Sprawdź które są nowe
                order_ids = [order['order_id'] for order in orders]
                existing_ids = service._get_existing_order_ids(order_ids)
                new_orders_count = len(order_ids) - len(existing_ids)
                has_new_orders = new_orders_count > 0
                
                reports_logger.info("Sprawdzenie nowych zamówień na stronie głównej",
                                  user_email=user_email,
                                  total_orders_from_baselinker=len(orders),
                                  existing_in_database=len(existing_ids),
                                  new_orders_count=new_orders_count,
                                  has_new_orders=has_new_orders,
                                  api_limit_reached=len(orders) >= 100)
                
                # UWAGA: Jeśli API zwrócił 100 zamówień, może być więcej
                if len(orders) >= 100:
                    reports_logger.warning("Limit API Baselinker osiągnięty na stronie głównej (4 dni)",
                                         user_email=user_email,
                                         orders_returned=len(orders),
                                         api_limit=100,
                                         info="Może być więcej nowych zamówień z ostatnich 4 dni niż pokazane")
            else:
                has_new_orders = False
                new_orders_count = 0
                reports_logger.info("Brak zamówień w Baselinker", user_email=user_email)
                
        except Exception as e:
            # POPRAWKA: Nie resetuj has_new_orders w przypadku błędu logowania
            reports_logger.error("Błąd sprawdzania nowych zamówień na stronie głównej",
                               user_email=user_email,
                               error=str(e),
                               error_type=type(e).__name__)
            
            # ZMIANA: W przypadku błędu, spróbuj prostszą metodę
            try:
                # Pobierz ostatnie 24h jako fallback
                date_from_fallback = datetime.now() - timedelta(hours=24)
                orders_fallback = service.fetch_orders_from_baselinker(date_from=date_from_fallback)
                
                if orders_fallback:
                    order_ids = [order['order_id'] for order in orders_fallback]
                    existing_ids = service._get_existing_order_ids(order_ids)
                    new_orders_count = len(order_ids) - len(existing_ids)
                    has_new_orders = new_orders_count > 0
                    
                    reports_logger.info("Fallback: Sprawdzenie zamówień z ostatnich 4 dni",
                                      user_email=user_email,
                                      fallback_orders=len(orders_fallback),
                                      new_orders_count=new_orders_count,
                                      has_new_orders=has_new_orders)
                else:
                    has_new_orders = False
                    new_orders_count = 0
                    
            except Exception as fallback_error:
                reports_logger.error("Błąd fallback sprawdzania zamówień",
                                   user_email=user_email,
                                   error=str(fallback_error))
                has_new_orders = False
                new_orders_count = 0
        
        # Pobierz podstawowe statystyki - ZMIANA: Domyślnie wszystkie dane
        date_from = None  # datetime.now().date() - timedelta(days=30)
        date_to = None    # datetime.now().date()
        
        reports_logger.info("Ładowanie statystyk Reports",
                          user_email=user_email,
                          date_from=date_from.isoformat() if date_from else "wszystkie",
                          date_to=date_to.isoformat() if date_to else "wszystkie")
        
        # POPRAWKA: Sprawdź czy są jakiekolwiek dane w bazie
        total_records = BaselinkerReportOrder.query.count()
        
        if total_records == 0:
            # Brak danych w bazie - ustaw puste statystyki
            stats = {
                'total_m3': 0.0,
                'order_amount_net': 0.0,
                'value_net': 0.0,
                'value_gross': 0.0,
                'avg_price_per_m3': 0.0,
                'delivery_cost': 0.0,
                'paid_amount_net': 0.0,
                'balance_due': 0.0,
                'production_volume': 0.0,
                'production_value_net': 0.0,
                'ready_pickup_volume': 0.0,
                'ready_pickup_value_net': 0.0
            }
            comparison = {}
            reports_logger.info("Brak danych w bazie - wyświetlanie pustych statystyk",
                              user_email=user_email)
        else:
            # Pobierz dane dla całej bazy lub wybranego zakresu dat
            query = BaselinkerReportOrder.get_filtered_orders(
                date_from=date_from,
                date_to=date_to
            )
            
            # Sprawdź czy query zwraca jakieś dane
            query_results = query.all()
            
            if not query_results:
                # Brak danych dla wybranego zakresu - ustaw puste statystyki
                stats = {
                    'total_m3': 0.0,
                    'order_amount_net': 0.0,
                    'value_net': 0.0,
                    'value_gross': 0.0,
                    'avg_price_per_m3': 0.0,
                    'delivery_cost': 0.0,
                    'paid_amount_net': 0.0,
                    'balance_due': 0.0,
                    'production_volume': 0.0,
                    'production_value_net': 0.0,
                    'ready_pickup_volume': 0.0,
                    'ready_pickup_value_net': 0.0
                }
                comparison = {}
                reports_logger.info("Brak danych dla wybranego zakresu",
                                  user_email=user_email,
                                  total_records_in_db=total_records)
            else:
                # Oblicz statystyki dla istniejących danych
                stats = BaselinkerReportOrder.get_statistics(query)
                
                # Oblicz porównania tylko jeśli mamy dane i sensowne wartości
                comparison = {}
                if stats.get('total_m3', 0) > 0 or stats.get('order_amount_net', 0) > 0:
                    try:
                        comparison = BaselinkerReportOrder.get_comparison_statistics(
                            {}, date_from, date_to
                        )
                    except Exception as comp_error:
                        reports_logger.warning("Błąd obliczania porównań", 
                                             user_email=user_email,
                                             error=str(comp_error))
                        comparison = {}
                
                reports_logger.info("Obliczono statystyki",
                                  user_email=user_email,
                                  records_count=len(query_results),
                                  total_m3=stats.get('total_m3', 0),
                                  order_amount_net=stats.get('order_amount_net', 0))
        
        # Pobierz ostatni log synchronizacji
        last_sync = ReportsSyncLog.query.order_by(ReportsSyncLog.sync_date.desc()).first()
        
        # Domyślne daty do wyświetlenia w interfejsie
        # Start: pierwszy dzień bieżącego miesiąca (lub poprzedniego, jeśli dziś jest 1.)
        today = datetime.now().date()
        first_day_current = today.replace(day=1)
        if today.day == 1:
            # Jeśli nowy miesiąc, sięgnij do początku poprzedniego
            prev_month_last_day = first_day_current - timedelta(days=1)
            default_date_from = prev_month_last_day.replace(day=1)
        else:
            default_date_from = first_day_current
        default_date_to = today
        
        context = {
            'user_email': user_email,
            'has_new_orders': has_new_orders,
            'new_orders_count': new_orders_count,
            'api_limit_reached': len(orders) >= 100 if 'orders' in locals() else False,  # NOWE
            'stats': stats,
            'comparison': comparison,
            'last_sync': last_sync,
            'default_date_from': default_date_from.isoformat(),
            'default_date_to': default_date_to.isoformat(),
            'total_records': total_records
        }
        
        return render_template('reports.html', **context)
        
    except Exception as e:
        reports_logger.error("Błąd ładowania strony Reports", 
                           user_email=user_email, 
                           error=str(e),
                           error_type=type(e).__name__)
        flash("Wystąpił błąd podczas ładowania danych.")
        return redirect(url_for('home'))

@reports_bp.route('/api/data')
@login_required
def api_get_data():
    """
    API endpoint do pobierania danych tabeli z filtrami
    """
    try:
        # Pobierz parametry filtrów
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        
        # Parsuj daty
        date_from = None
        date_to = None
        
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'error': 'Nieprawidłowy format daty początkowej'}), 400
                
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'error': 'Nieprawidłowy format daty końcowej'}), 400
        
        # Pobierz filtry kolumn (obsługa multiple values)
        column_filters = {}
        for key, values in request.args.items(multi=True):
            if key.startswith('filter_') and values:
                column_name = key.replace('filter_', '')
                if hasattr(BaselinkerReportOrder, column_name):
                    # Grupuj wartości dla tego samego filtra
                    if column_name not in column_filters:
                        column_filters[column_name] = []
                    column_filters[column_name].extend(values if isinstance(values, list) else [values])
        
        # Wyczyść puste wartości
        for column_name in list(column_filters.keys()):
            column_filters[column_name] = [v for v in column_filters[column_name] if v and v.strip()]
            if not column_filters[column_name]:
                del column_filters[column_name]
        
        reports_logger.debug("Pobieranie danych z filtrami",
                           date_from=date_from.isoformat() if date_from else None,
                           date_to=date_to.isoformat() if date_to else None,
                           filters=column_filters)
        
        try:
            # Pobierz dane z bazy
            query = BaselinkerReportOrder.get_filtered_orders(
                filters=column_filters,
                date_from=date_from,
                date_to=date_to
            )

            orders = query.all()

            # DEBUG: Dodane rozszerzone logowanie
            manual_orders = [o for o in orders if o.is_manual]
            all_manual_in_db = BaselinkerReportOrder.query.filter(BaselinkerReportOrder.is_manual == True).all()
            
        except Exception as db_error:
            # Jeśli błąd bazy danych, spróbuj ponownie
            reports_logger.warning("Błąd bazy danych, ponowna próba", error=str(db_error))
            try:
                # Zamknij połączenie i spróbuj ponownie
                db.session.close()
                db.engine.dispose()
                
                query = BaselinkerReportOrder.get_filtered_orders(
                    filters=column_filters,
                    date_from=date_from,
                    date_to=date_to
                )
                orders = query.all()
                
            except Exception as retry_error:
                reports_logger.error("Błąd bazy danych przy ponownej próbie", error=str(retry_error))
                return jsonify({
                    'success': False,
                    'error': f'Błąd bazy danych: {str(retry_error)}'
                }), 500
        
        # Konwertuj na JSON
        data = [order.to_dict() for order in orders]
        
        # Oblicz statystyki dla przefiltrowanych danych
        stats = BaselinkerReportOrder.get_statistics(query)
        
        # Oblicz statystyki porównawcze jeśli mamy daty
        comparison = {}
        if date_from and date_to:
            try:
                comparison = BaselinkerReportOrder.get_comparison_statistics(
                    column_filters, date_from, date_to
                )
            except Exception as comp_error:
                reports_logger.warning("Błąd obliczania porównań", error=str(comp_error))
                comparison = {}
        
        return jsonify({
            'success': True,
            'data': data,
            'stats': stats,
            'comparison': comparison,
            'total_count': len(data)
        })
        
    except Exception as e:
        reports_logger.error("Błąd pobierania danych API", error=str(e))
        return jsonify({
            'success': False,
            'error': f'Błąd serwera: {str(e)}'
        }), 500


@reports_bp.route('/api/sync', methods=['POST'])
@login_required
def api_sync_with_baselinker():
    """
    API endpoint do synchronizacji z Baselinker
    """
    user_email = session.get('user_email')
    
    try:
        # Pobierz parametry
        data = request.get_json() or {}
        date_from_str = data.get('date_from')
        date_to_str = data.get('date_to')
        selected_orders = data.get('selected_orders', [])  # Lista ID zamówień do synchronizacji
        
        # ZMIANA: Parsuj daty tylko jeśli podane, nie używaj domyślnie 30 dni
        date_from = None
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
                reports_logger.info("Użyto podanej daty rozpoczęcia", 
                                  date_from=date_from.isoformat())
            except ValueError as e:
                reports_logger.warning("Błędny format daty, pominięto filtr daty",
                                     date_from_str=date_from_str,
                                     error=str(e))
                date_from = None
        
        reports_logger.info("Rozpoczęcie synchronizacji z Baselinker",
                          user_email=user_email,
                          date_from=date_from.isoformat() if date_from else "wszystkie zamówienia",
                          selected_orders_count=len(selected_orders),
                          selected_orders=selected_orders[:10])  # Loguj tylko pierwsze 10 ID
        
        # Pobierz serwis
        service = get_reports_service()
        
        if selected_orders and len(selected_orders) > 0:
            # Synchronizuj wybrane zamówienia
            reports_logger.info("Synchronizowanie wybranych zamówień", 
                              selected_orders_count=len(selected_orders))
            result = _sync_selected_orders(service, selected_orders)
        else:
            # ZMIANA: Synchronizuj wszystkie zamówienia (bez ograniczenia 30 dni)
            reports_logger.info("Synchronizowanie wszystkich zamówień",
                              has_date_filter=date_from is not None)
            result = service.sync_orders(date_from=date_from, sync_type='manual')
        
        reports_logger.info("Synchronizacja zakończona",
                          user_email=user_email,
                          success=result.get('success'),
                          orders_processed=result.get('orders_processed', 0),
                          orders_added=result.get('orders_added', 0),
                          orders_updated=result.get('orders_updated', 0))
        
        return jsonify(result)
        
    except Exception as e:
        reports_logger.error("Błąd synchronizacji",
                           user_email=user_email,
                           error=str(e),
                           error_type=type(e).__name__)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@reports_bp.route('/api/add-manual-row', methods=['POST'])
@login_required
def api_add_manual_row():
    """
    ZAKTUALIZOWANY: API endpoint do dodawania ręcznego wiersza z obsługą produktów
    """
    user_email = session.get('user_email')
    
    try:
        data = request.get_json()
        
        reports_logger.info("Dodawanie ręcznego wiersza",
                          user_email=user_email,
                          data_keys=list(data.keys()) if data else None)
        
        # DEBUG: Wyloguj otrzymane dane
        reports_logger.debug("Otrzymane dane formularza",
                           user_email=user_email,
                           data=data)
        
        # NOWA LOGIKA: Obsługa produktów z nowej struktury
        products_data = data.get('products', [])
        
        if not products_data:
            # FALLBACK: Jeśli brak produktów, spróbuj stara struktura
            reports_logger.warning("Brak produktów w nowej strukturze, używam fallback",
                                 user_email=user_email)
            
            # Utwórz pojedynczy rekord (stara metoda)
            record = BaselinkerReportOrder(
                is_manual=True,
                date_created=datetime.strptime(data.get('date_created'), '%Y-%m-%d').date() if data.get('date_created') else date.today(),
                internal_order_number=data.get('internal_order_number'),
                customer_name=data.get('customer_name'),
                delivery_postcode=data.get('delivery_postcode'),
                delivery_city=data.get('delivery_city'),
                delivery_address=data.get('delivery_address'),
                delivery_state=data.get('delivery_state'),
                phone=data.get('phone'),
                caretaker=user_email,
                delivery_method=data.get('delivery_method'),
                order_source=data.get('order_source'),
                group_type=data.get('group_type'),
                product_type=data.get('product_type', 'klejonka'),
                finish_state=data.get('finish_state', 'surowy'),
                wood_species=data.get('wood_species'),
                technology=data.get('technology'),
                wood_class=data.get('wood_class'),
                length_cm=float(data.get('length_cm', 0)) if data.get('length_cm') else None,
                width_cm=float(data.get('width_cm', 0)) if data.get('width_cm') else None,
                thickness_cm=float(data.get('thickness_cm', 0)) if data.get('thickness_cm') else None,
                quantity=int(data.get('quantity', 1)) if data.get('quantity') else 1,
                # ZMIANA: Obsługa price_net zamiast price_gross
                price_net=float(data.get('price_net', 0)) if data.get('price_net') else None,
                price_gross=float(data.get('price_net', 0)) * 1.23 if data.get('price_net') else None,
                delivery_cost=float(data.get('delivery_cost', 0)) if data.get('delivery_cost') else None,
                payment_method=data.get('payment_method'),
                paid_amount_net=float(data.get('paid_amount_net', 0)) if data.get('paid_amount_net') else 0,
                current_status=data.get('current_status', 'Nowe - opłacone'),
                order_amount_net=float(data.get('price_net', 0)) * int(data.get('quantity', 1)) if data.get('price_net') else 0
            )
            
            # Oblicz pola pochodne
            record.calculate_fields()
            
            # Zapisz do bazy
            db.session.add(record)
            db.session.commit()

            # DEBUG: Sprawdź zapisany rekord
            created_records = [record]

            # Sprawdź czy rekordy są w bazie
            fresh_records = BaselinkerReportOrder.query.filter(
                BaselinkerReportOrder.id.in_([r.id for r in created_records])
            ).all()
            
            reports_logger.info("Dodano ręczny wiersz (fallback)",
                              user_email=user_email,
                              record_id=record.id)
            
            return jsonify({
                'success': True,
                'message': 'Wiersz został dodany',
                'record_id': record.id
            })
            
        else:
            # NOWA LOGIKA: Obsługa wielu produktów
            reports_logger.info("Dodawanie zamówienia z produktami",
                              user_email=user_email,
                              products_count=len(products_data))
            
            created_records = []
            
            for i, product_data in enumerate(products_data):
                reports_logger.debug(f"Przetwarzanie produktu {i+1}",
                                   user_email=user_email,
                                   product_data=product_data)
                
                # Utwórz rekord dla każdego produktu
                record = BaselinkerReportOrder(
                    is_manual=True,
                    
                    # Wspólne dane zamówienia
                    date_created=datetime.strptime(data.get('date_created'), '%Y-%m-%d').date() if data.get('date_created') else date.today(),
                    internal_order_number=data.get('internal_order_number'),
                    customer_name=data.get('customer_name'),
                    delivery_postcode=data.get('delivery_postcode'),
                    delivery_city=data.get('delivery_city'),
                    delivery_address=data.get('delivery_address'),
                    delivery_state=data.get('delivery_state'),
                    phone=data.get('phone'),
                    caretaker=user_email,
                    delivery_method=data.get('delivery_method'),
                    order_source=data.get('order_source'),
                    delivery_cost=float(data.get('delivery_cost', 0)) if data.get('delivery_cost') else None,
                    payment_method=data.get('payment_method'),
                    paid_amount_net=float(data.get('paid_amount_net', 0)) if data.get('paid_amount_net') else 0,
                    current_status=data.get('current_status', 'Nowe - opłacone'),
                    
                    # Dane produktu
                    group_type=product_data.get('group_type'),
                    product_type=product_data.get('product_type'),
                    finish_state=product_data.get('finish_state', 'surowy'),
                    wood_species=product_data.get('wood_species'),
                    technology=product_data.get('technology'),
                    wood_class=product_data.get('wood_class'),
                    length_cm=float(product_data.get('length_cm', 0)) if product_data.get('length_cm') else None,
                    width_cm=float(product_data.get('width_cm', 0)) if product_data.get('width_cm') else None,
                    thickness_cm=float(product_data.get('thickness_cm', 0)) if product_data.get('thickness_cm') else None,
                    quantity=int(product_data.get('quantity', 1)) if product_data.get('quantity') else 1,
                    
                    # ZMIANA: Obsługa price_net z produktu
                    price_net=float(product_data.get('price_net', 0)) if product_data.get('price_net') else None,
                    price_gross=float(product_data.get('price_net', 0)) * 1.23 if product_data.get('price_net') else None,
                    order_amount_net=float(product_data.get('price_net', 0)) * int(product_data.get('quantity', 1)) if product_data.get('price_net') else 0
                )
                
                # Oblicz pola pochodne
                record.calculate_fields()
                
                # Dodaj do sesji
                db.session.add(record)
                created_records.append(record)
            
            # Zapisz wszystkie rekordy
            db.session.commit()

            # Sprawdź czy rekordy są w bazie
            fresh_records = BaselinkerReportOrder.query.filter(
                BaselinkerReportOrder.id.in_([r.id for r in created_records])
            ).all()
            
            # Pobierz ID wszystkich utworzonych rekordów
            record_ids = [record.id for record in created_records]
            
            reports_logger.info("Dodano zamówienie wieloproduktowe",
                              user_email=user_email,
                              products_count=len(created_records),
                              record_ids=record_ids)
            
            return jsonify({
                'success': True,
                'message': f'Zamówienie z {len(created_records)} produktami zostało dodane',
                'record_ids': record_ids,
                'products_count': len(created_records)
            })
        
    except Exception as e:
        db.session.rollback()
        reports_logger.error("Błąd dodawania ręcznego wiersza",
                           user_email=user_email,
                           error=str(e),
                           error_type=type(e).__name__)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@reports_bp.route('/api/update-manual-row', methods=['POST'])
@login_required
def api_update_manual_row():
    """
    ZAKTUALIZOWANY: API endpoint do edycji rekordów z obsługą wielu produktów
    """
    user_email = session.get('user_email')
    
    try:
        data = request.get_json()
        record_id = data.get('record_id')
        
        if not record_id:
            return jsonify({'success': False, 'error': 'Brak ID rekordu'}), 400
        
        # Pobierz główny rekord
        main_record = BaselinkerReportOrder.query.get_or_404(record_id)
        
        record_type = "ręczny" if main_record.is_manual else "z Baselinker"
        reports_logger.info(f"Edycja rekordu ({record_type})",
                          user_email=user_email,
                          record_id=record_id)
        
        # NOWA LOGIKA: Obsługa wielu produktów
        products_data = data.get('products', [])

        if products_data and len(products_data) > 1:
            # WIELOPRODUKTOWE ZAMÓWIENIE (głównie z Baselinker)
            reports_logger.info(f"Aktualizacja zamówienia wieloproduktowego",
                              user_email=user_email,
                              record_id=record_id,
                              products_count=len(products_data))
            
            # Pobierz wszystkie rekordy tego zamówienia
            if main_record.baselinker_order_id:
                all_order_records = BaselinkerReportOrder.query.filter_by(
                    baselinker_order_id=main_record.baselinker_order_id
                ).all()
            else:
                all_order_records = [main_record]
            
            # Utwórz mapę istniejących rekordów po ID
            existing_records_map = {record.id: record for record in all_order_records}
            
            updated_records = []
            
            for product_data in products_data:
                product_record_id = product_data.get('record_id')
                
                if product_record_id and product_record_id in existing_records_map:
                    # Aktualizuj istniejący rekord produktu
                    record = existing_records_map[product_record_id]
                    
                    # Aktualizuj wspólne dane zamówienia (z pierwszego produktu)
                    if record == main_record:
                        _update_order_common_fields(record, data)
                    
                    # Aktualizuj dane produktu
                    _update_product_fields(record, product_data)
                    
                    updated_records.append(record)
                    
                else:
                    # TO-DO: Obsługa nowych produktów (jeśli będzie potrzebna)
                    reports_logger.warning("Próba dodania nowego produktu do istniejącego zamówienia",
                                         user_email=user_email,
                                         main_record_id=record_id,
                                         product_data=product_data)
            
            # Przelicz pola dla wszystkich zaktualizowanych rekordów
            for record in updated_records:
                record.calculate_fields()
                record.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            reports_logger.info("Zaktualizowano zamówienie wieloproduktowe",
                              user_email=user_email,
                              main_record_id=record_id,
                              updated_products_count=len(updated_records))
            
            return jsonify({
                'success': True,
                'message': f'Zamówienie z {len(updated_records)} produktami zostało zaktualizowane',
                'updated_products': len(updated_records)
            })
            
        else:
            # JEDNOPRODUKTOWE ZAMÓWIENIE (ręczne lub pojedynczy produkt z Baselinker)
            _update_order_common_fields(main_record, data)
            
            if products_data and len(products_data) > 0:
                _update_product_fields(main_record, products_data[0])
            
            # Przelicz pola
            main_record.calculate_fields()
            main_record.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Rekord został zaktualizowany',
                'record_id': main_record.id
            })
        
    except Exception as e:
        db.session.rollback()
        reports_logger.error("Błąd aktualizacji rekordu",
                           user_email=user_email,
                           record_id=record_id,
                           error=str(e))
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

def _update_order_common_fields(record, data):
    """
    Pomocnicza funkcja do aktualizacji wspólnych pól zamówienia
    """
    common_fields = [
        'date_created', 'baselinker_order_id', 'internal_order_number', 'customer_name', 
        'delivery_postcode', 'delivery_city', 'delivery_address', 'delivery_state',
        'phone', 'caretaker', 'delivery_method', 'order_source',
        'delivery_cost', 'payment_method', 'paid_amount_net', 'current_status'
    ]
    
    for field in common_fields:
        if field in data and data[field] is not None:
            value = data[field]
            if field == 'date_created' and value:
                setattr(record, field, datetime.strptime(value, '%Y-%m-%d').date())
            elif field in ['delivery_cost', 'paid_amount_net'] and value:
                setattr(record, field, float(value))
            else:
                setattr(record, field, value)

def _update_product_fields(record, product_data):
    """
    Pomocnicza funkcja do aktualizacji pól produktu
    """
    product_fields = {
        'group_type': str,
        'product_type': str,
        'wood_species': str,
        'technology': str,
        'wood_class': str,
        'finish_state': str,
        'length_cm': float,
        'width_cm': float,
        'thickness_cm': float,
        'quantity': int,
        'price_net': float,
        # NOWE POLA:
        'price_type': str,
        'original_amount_from_baselinker': float
    }
    
    for field, field_type in product_fields.items():
        if field in product_data and product_data[field] is not None:
            value = product_data[field]
            
            if field == 'price_net' and value:
                # Konwersja price_net na price_gross
                price_net = float(value)
                setattr(record, 'price_net', price_net)

                price_gross = round(price_net * 1.23, 2)
                setattr(record, 'price_gross', price_gross)

                # Ustaw order_amount_net na podstawie price_net i quantity
                quantity = int(product_data.get('quantity', record.quantity or 1))
                order_amount_net = round(price_net * quantity, 2)
                setattr(record, 'order_amount_net', order_amount_net)

                # Dodatkowo aktualizuj wartości netto/brutto produktu
                setattr(record, 'value_net', order_amount_net)
                setattr(record, 'value_gross', round(price_gross * quantity, 2))
            elif field == 'original_amount_from_baselinker' and value:
                # NOWE: Obsługa oryginalnej kwoty z Baselinker
                setattr(record, field, float(value))
            elif field == 'price_type' and value:
                # NOWE: Obsługa typu ceny
                # Normalizuj wartość
                normalized_value = value.strip().lower()
                if normalized_value in ['netto', 'brutto', '']:
                    setattr(record, field, normalized_value)
                else:
                    # Jeśli nieznana wartość, ustaw jako puste
                    setattr(record, field, '')
            elif field_type == float and value:
                setattr(record, field, float(value))
            elif field_type == int and value:
                setattr(record, field, int(value))
            elif field_type == str:
                setattr(record, field, str(value) if value else None)

@reports_bp.route('/api/export-excel')
@login_required
def api_export_excel():
    """
    API endpoint do eksportu danych do Excel z zaawansowanym formatowaniem
    POPRAWKA: Pełna obsługa kodowania UTF-8 i zabezpieczenia przed błędami
    """
    user_email = session.get('user_email')
    
    try:
        # POPRAWKA: Wymuś kodowanie UTF-8 dla całej funkcji
        import sys
        import os
        if hasattr(sys.stdout, 'reconfigure'):
            try:
                sys.stdout.reconfigure(encoding='utf-8')
                sys.stderr.reconfigure(encoding='utf-8')
            except:
                pass
        os.environ['PYTHONIOENCODING'] = 'utf-8'
        
        # Pobierz parametry filtrów
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        
        # Parsuj daty
        date_from = None
        date_to = None
        
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except ValueError:
                pass
                
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Pobierz filtry kolumn
        column_filters = {}
        for key, values in request.args.items(multi=True):
            if key.startswith('filter_') and values:
                column_name = key.replace('filter_', '')
                if hasattr(BaselinkerReportOrder, column_name):
                    if column_name not in column_filters:
                        column_filters[column_name] = []
                    column_filters[column_name].extend(values if isinstance(values, list) else [values])
        
        # Wyczyść puste wartości
        for column_name in list(column_filters.keys()):
            column_filters[column_name] = [v for v in column_filters[column_name] if v and v.strip()]
            if not column_filters[column_name]:
                del column_filters[column_name]
        
        reports_logger.info("Eksport do Excel",
                          user_email=user_email,
                          date_from=date_from.isoformat() if date_from else None,
                          date_to=date_to.isoformat() if date_to else None,
                          filters_count=len(column_filters))
        
        # Pobierz dane
        query = BaselinkerReportOrder.get_filtered_orders(
            filters=column_filters,
            date_from=date_from,
            date_to=date_to
        )
        orders = query.all()
        
        if not orders:
            return jsonify({
                'success': False,
                'error': 'Brak danych do eksportu'
            }), 400
        
        # POPRAWKA: Agresywna funkcja do obsługi polskich znaków
        def safe_str(value):
            if value is None:
                return ''
            try:
                # Najpierw spróbuj normalnie
                result = str(value)
                # Usuń/zastąp problematyczne polskie znaki prewencyjnie
                result = result.replace('ó', 'o').replace('ą', 'a').replace('ć', 'c')
                result = result.replace('ę', 'e').replace('ł', 'l').replace('ń', 'n')
                result = result.replace('ś', 's').replace('ź', 'z').replace('ż', 'z')
                result = result.replace('Ó', 'O').replace('Ą', 'A').replace('Ć', 'C')
                result = result.replace('Ę', 'E').replace('Ł', 'L').replace('Ń', 'N')
                result = result.replace('Ś', 'S').replace('Ź', 'Z').replace('Ż', 'Z')
                return result
            except UnicodeEncodeError:
                # Jeśli błąd, zastąp problematyczne znaki
                return str(value).encode('ascii', errors='replace').decode('ascii')
            except Exception:
                # Ostateczność - usuń wszystkie nie-ASCII
                return ''.join(char for char in str(value) if ord(char) < 128)
        
        # Utwórz plik Excel w pamięci
        output = io.BytesIO()
        
        # Import dla stylow Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        # POPRAWKA: Ustaw kodowanie dla polskich znaków
        import locale
        try:
            locale.setlocale(locale.LC_ALL, 'pl_PL.UTF-8')
        except:
            try:
                locale.setlocale(locale.LC_ALL, 'Polish_Poland.1250')
            except:
                try:
                    locale.setlocale(locale.LC_ALL, 'C.UTF-8')
                except:
                    pass  # Użyj domyślnego kodowania
        
        # Utwórz workbook
        workbook = Workbook()
        
        # ===== ARKUSZ 1: DANE SZCZEGÓŁOWE =====
        ws_details = workbook.active
        ws_details.title = "Dane szczegolowe"  # Bez polskich znaków w tytule

        # NOWE: Oblicz TTL m³ dla każdego zamówienia I ŚLEDŹ PIERWSZY PRODUKT
        order_volumes = {}
        order_first_product = {}  # Śledzi pierwszy produkt w każdym zamówieniu
        
        for idx, order in enumerate(orders):
            # Identyfikator zamówienia (Baselinker ID lub manual ID)
            order_key = order.baselinker_order_id or f"manual_{order.id}"

            if order_key not in order_volumes:
                order_volumes[order_key] = 0.0
                order_first_product[order_key] = idx  # Zapisz indeks pierwszego produktu

            # Dodaj objętość tego produktu do sumy zamówienia
            order_volumes[order_key] += float(order.total_volume or 0)

        # Przygotuj dane do DataFrame
        excel_data = []
        for idx, order in enumerate(orders):
            # KLUCZOWA POPRAWKA: Kolumny poziomu zamówienia tylko dla pierwszego produktu
            order_key = order.baselinker_order_id or f"manual_{order.id}"
            is_first_product_in_order = order_first_product.get(order_key) == idx
            
            # Wartości pokazywane tylko raz na zamówienie (w pierwszym produkcie)
            if is_first_product_in_order:
                calculated_ttl_m3 = order_volumes.get(order_key, 0.0)
                kwota_zamowien_netto = float(order.order_amount_net or 0)
                nr_baselinker = safe_str(order.baselinker_order_id)
                nr_wew = safe_str(order.internal_order_number)
                nazwa_klienta = safe_str(order.customer_name)
                kod_pocztowy = safe_str(order.delivery_postcode)
                miejscowosc = safe_str(order.delivery_city)
                ulica = safe_str(order.delivery_address)
                wojewodztwo = safe_str(order.delivery_state)
                telefon = safe_str(order.phone)
                opiekun = safe_str(order.caretaker)
                dostawa = safe_str(order.delivery_method)
                zrodlo = safe_str(order.order_source)
                koszt_kuriera = float(order.delivery_cost or 0)
                koszt_dostawy_netto = float(order.delivery_cost or 0) / 1.23
                sposob_platnosci = safe_str(order.payment_method)
                zaplacono_netto = float(order.paid_amount_net or 0)
                do_zaplaty_netto = float(order.balance_due or 0)
                srednia_cena_za_m3 = float(order.avg_order_price_per_m3 or 0)
            else:
                # Pozostałe produkty w zamówieniu mają 0/pusty string dla kolumn poziomu zamówienia
                calculated_ttl_m3 = 0.0
                kwota_zamowien_netto = 0.0
                nr_baselinker = ''
                nr_wew = ''
                nazwa_klienta = ''
                kod_pocztowy = ''
                miejscowosc = ''
                ulica = ''
                wojewodztwo = ''
                telefon = ''
                opiekun = ''
                dostawa = ''
                zrodlo = ''
                koszt_kuriera = 0.0
                koszt_dostawy_netto = 0.0
                sposob_platnosci = ''
                zaplacono_netto = 0.0
                do_zaplaty_netto = 0.0
                srednia_cena_za_m3 = 0.0
    
            excel_data.append({
                # KOLUMNY POZIOMU ZAMÓWIENIA (tylko w pierwszym produkcie)
                'Data': order.date_created.strftime('%d-%m-%Y') if order.date_created else '',
                'TTL m3': calculated_ttl_m3,  # POPRAWIONE: używa zmiennej
                'Kwota zamowien netto': kwota_zamowien_netto,  # POPRAWIONE: używa zmiennej
                'Nr Baselinker': nr_baselinker,  # POPRAWIONE: używa zmiennej
                'Nr wew.': nr_wew,  # POPRAWIONE: używa zmiennej
                'Nazwa klienta': nazwa_klienta,  # POPRAWIONE: używa zmiennej
                'Kod pocztowy': kod_pocztowy,  # POPRAWIONE: używa zmiennej
                'Miejscowosc': miejscowosc,  # POPRAWIONE: używa zmiennej
                'Ulica': ulica,  # POPRAWIONE: używa zmiennej
                'Wojewodztwo': wojewodztwo,  # POPRAWIONE: używa zmiennej
                'Telefon': telefon,  # POPRAWIONE: używa zmiennej
                'Opiekun': opiekun,  # POPRAWIONE: używa zmiennej
                'Dostawa': dostawa,  # POPRAWIONE: używa zmiennej
                'Zrodlo': zrodlo,  # POPRAWIONE: używa zmiennej
                
                # KOLUMNY POZIOMU PRODUKTU (zawsze pokazywane)
                'Grupa': safe_str(order.group_type),
                'Rodzaj': safe_str(order.product_type),
                'Wykonczenie': safe_str(order.finish_state),
                'Gatunek': safe_str(order.wood_species),
                'Technologia': safe_str(order.technology),
                'Klasa': safe_str(order.wood_class),
                'Dlugosc': float(order.length_cm or 0),
                'Szerokosc': float(order.width_cm or 0),
                'Grubosc': float(order.thickness_cm or 0),
                'Ilosc': int(order.quantity or 0),
                'Cena brutto': float(order.price_gross or 0),
                'Cena netto': float(order.price_net or 0),
                'Wartosc brutto': float(order.value_gross or 0),
                'Wartosc netto': float(order.value_net or 0),
                'Objetosc 1 szt.': float(order.volume_per_piece or 0),
                'Objetosc TTL': float(order.total_volume or 0),
                'Cena za m3': float(order.price_per_m3 or 0),
                'Srednia cena za m3': srednia_cena_za_m3,
                'Data realizacji': order.realization_date.strftime('%d-%m-%Y') if order.realization_date else '',
                'Status': safe_str(order.current_status),
                
                # KOLUMNY FINANSOWE POZIOMU ZAMÓWIENIA (tylko w pierwszym produkcie)
                'Koszt kuriera': koszt_kuriera,  # POPRAWIONE: używa zmiennej
                'Koszt dostawy netto': koszt_dostawy_netto,  # POPRAWIONE: używa zmiennej
                'Sposob platnosci': sposob_platnosci,  # POPRAWIONE: używa zmiennej
                'Zaplacono netto': zaplacono_netto,  # POPRAWIONE: używa zmiennej
                'Do zaplaty netto': do_zaplaty_netto,  # POPRAWIONE: używa zmiennej
                
                # KOLUMNY PRODUKCJI (poziom produktu - zawsze pokazywane)
                'Ilosc w produkcji': float(order.production_volume or 0),
                'Wartosc w produkcji': float(order.production_value_net or 0),
                'Wyprodukowano': float(order.ready_pickup_volume or 0),  # NOWA KOLUMNA
                'Wartosc wyprodukowana netto': float(order.ready_pickup_value_net or 0),
                'Gotowe do odbioru': float(0.0)  # Dostosuj logikę według potrzeb
            })
        
        # POPRAWKA: Sprawdź czy excel_data nie jest puste
        if not excel_data:
            return jsonify({
                'success': False,
                'error': 'Brak danych do eksportu po przetworzeniu filtrów'
            }), 400

        # POPRAWKA: Dodatkowa normalizacja wszystkich stringów przed DataFrame
        for row in excel_data:
            for key, value in row.items():
                if isinstance(value, str) and value:
                    # Dodatkowe czyszczenie dla pewności
                    try:
                        # Zamień wszelkie pozostałe problematyczne znaki
                        value = value.encode('ascii', errors='ignore').decode('ascii')
                        row[key] = value
                    except:
                        row[key] = ''

        # Utwórz DataFrame
        df = pd.DataFrame(excel_data)
        
        # Sprawdź czy DataFrame ma dane
        if df.empty or len(df) == 0:
            return jsonify({
                'success': False,
                'error': 'DataFrame jest pusty po konwersji'
            }), 400
        
        # ===== DEFINICJA KOLORÓW (PASTELOWE) =====
        COLORS = {
            'order_data': 'E3F2FD',      # Jasny niebieski
            'customer_data': 'E8F5E8',   # Jasny zielony  
            'logistics_data': 'FFF9C4',  # Jasny żółty
            'product_data': 'FFE0B2',    # Jasny pomarańczowy
            'financial_data': 'F3E5F5',  # Jasny fioletowy
            'production_data': 'F5F5F5'  # Jasny szary
        }
        
        # Mapowanie kolumn do kolorów (bez polskich znaków)
        COLUMN_COLORS = {
            'Data': 'order_data',
            'TTL m3': 'order_data',
            'Kwota zamowien netto': 'order_data',
            'Nr Baselinker': 'order_data',
            'Nr wew.': 'order_data',
            'Nazwa klienta': 'customer_data',  # POPRAWIONE: było 'Imie i nazwisko'
            'Kod pocztowy': 'customer_data',
            'Miejscowosc': 'customer_data',
            'Ulica': 'customer_data',
            'Wojewodztwo': 'customer_data',
            'Telefon': 'customer_data',
            'Opiekun': 'customer_data',
            'Dostawa': 'logistics_data',
            'Zrodlo': 'logistics_data',
            'Status': 'logistics_data',
            'Sposob platnosci': 'logistics_data',
            'Grupa': 'product_data',
            'Rodzaj': 'product_data',
            'Wykonczenie': 'product_data',
            'Gatunek': 'product_data',
            'Technologia': 'product_data',
            'Klasa': 'product_data',
            'Dlugosc': 'product_data',
            'Szerokosc': 'product_data',
            'Grubosc': 'product_data',
            'Ilosc': 'product_data',
            'Cena brutto': 'financial_data',
            'Cena netto': 'financial_data',
            'Wartosc brutto': 'financial_data',
            'Wartosc netto': 'financial_data',
            'Cena za m3': 'financial_data',
            'Srednia cena za m3': 'financial_data',
            'Koszt kuriera': 'financial_data',
            'Koszt dostawy netto': 'financial_data',
            'Zaplacono netto': 'financial_data',
            'Do zaplaty netto': 'financial_data',
            'Objetosc 1 szt.': 'production_data',
            'Objetosc TTL': 'production_data',
            'Data realizacji': 'production_data',
            'Ilosc w produkcji': 'production_data',
            'Wartosc w produkcji': 'production_data',
            'Wyprodukowano': 'production_data',  # NOWA KOLUMNA
            'Gotowe do odbioru': 'production_data'
        }
        
        # Kolumny liczbowe dla formuł podsumowania (bez polskich znaków)
        NUMERIC_COLUMNS = {
            'TTL m3': 'SUM',
            'Kwota zamowien netto': 'SUM',
            'Dlugosc': 'AVERAGE',
            'Szerokosc': 'AVERAGE', 
            'Grubosc': 'AVERAGE',
            'Ilosc': 'SUM',
            'Cena brutto': 'SUM',
            'Cena netto': 'SUM',
            'Wartosc brutto': 'SUM',
            'Wartosc netto': 'SUM',
            'Objetosc 1 szt.': 'AVERAGE',
            'Objetosc TTL': 'SUM',
            'Cena za m3': 'AVERAGE',
            'Koszt kuriera': 'SUM',
            'Koszt dostawy netto': 'SUM',
            'Zaplacono netto': 'SUM',
            'Do zaplaty netto': 'SUM',
            'Ilosc w produkcji': 'SUM',
            'Wartosc w produkcji': 'SUM',
            'Wyprodukowano': 'SUM',  # NOWA KOLUMNA
            'Gotowe do odbioru': 'SUM'
        }
        
        # Stylizacja
        header_font = Font(bold=True, color='FFFFFF')
        summary_font = Font(bold=True, color='333333')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ===== FORMATOWANIE ARKUSZA GŁÓWNEGO =====
        def format_details_sheet(worksheet, dataframe):
            # POPRAWKA: Sprawdź czy DataFrame ma dane
            if len(dataframe) == 0:
                # Dodaj tylko nagłówek informujący o braku danych
                cell = worksheet.cell(row=1, column=1, value="Brak danych do wyswietlenia")
                cell.font = Font(bold=True, size=14, color='FF0000')
                return
    
            # Wyczyść arkusz - POPRAWKA: Sprawdź czy arkusz ma wiersze
            if worksheet.max_row > 1:
                worksheet.delete_rows(1, worksheet.max_row)

            # WIERSZ 1: NAGŁÓWKI
            headers = list(dataframe.columns)
            for col_idx, header in enumerate(headers, 1):
                try:
                    cell = worksheet.cell(row=1, column=col_idx, value=safe_str(header))
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                    
                    # Kolor tła nagłówka
                    color_key = COLUMN_COLORS.get(header, 'order_data')
                    header_colors = {
                        'order_data': '1976D2',
                        'customer_data': '388E3C',
                        'logistics_data': 'F57C00',
                        'product_data': 'F57C00',
                        'financial_data': '7B1FA2',
                        'production_data': '616161'
                    }
                    cell.fill = PatternFill(start_color=header_colors[color_key], end_color=header_colors[color_key], fill_type='solid')
                except Exception as e:
                    reports_logger.warning(f"Błąd tworzenia nagłówka kolumny {col_idx}: {e}")
                    continue
            
            # WIERSZ 2: PODSUMOWANIA (FORMUŁY)
            data_start_row = 4  # Dane zaczynają się od wiersza 4
            data_end_row = data_start_row + len(dataframe) - 1

            # POPRAWKA: Zabezpieczenie przed nieprawidłowymi zakresami
            if len(dataframe) == 0:
                return
            elif data_end_row < data_start_row:
                data_end_row = data_start_row
            
            for col_idx, header in enumerate(headers, 1):
                try:
                    cell = worksheet.cell(row=2, column=col_idx)
                    cell.font = summary_font
                    cell.border = border
                    
                    # Kolor tła podsumowania
                    color_key = COLUMN_COLORS.get(header, 'order_data')
                    cell.fill = PatternFill(start_color=COLORS[color_key], end_color=COLORS[color_key], fill_type='solid')
                    
                    # Dodaj formułę dla kolumn liczbowych - tylko jeśli są dane
                    if header in NUMERIC_COLUMNS and len(dataframe) > 0:
                        formula_type = NUMERIC_COLUMNS[header]
                        col_letter = get_column_letter(col_idx)
                        
                        # POPRAWKA: Specjalne formuły dla pól które mogą być duplikowane per zamówienie
                        scalable_fields = ['Kwota zamowien netto', 'Koszt kuriera', 'Koszt dostawy netto', 'Zaplacono netto', 'Saldo']
                        
                        if header in scalable_fields and formula_type == 'SUM':
                            # Użyj prostej SUM zamiast skomplikowanego SUMPRODUCT
                            cell.value = f'=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})'
                        elif formula_type == 'SUM':
                            cell.value = f'=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})'
                        elif formula_type == 'AVERAGE':
                            cell.value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})'
                        
                        # Format liczbowy (bez polskich znaków)
                        if 'zl' in header or 'Kwota' in header or 'Wartosc' in header or 'Cena' in header or 'Koszt' in header or 'Zaplacono' in header or 'Saldo' in header:
                            cell.number_format = '#,##0.00" zl"'
                        elif 'm3' in header or 'Objetosc' in header:
                            cell.number_format = '#,##0.0000'
                        else:
                            cell.number_format = '#,##0.00'
                    else:
                        # Dla kolumn tekstowych - informacje opisowe
                        if header == 'Data':
                            period_text = f"Okres: {date_from.strftime('%d-%m-%Y') if date_from else 'wszystkie'} - {date_to.strftime('%d-%m-%Y') if date_to else 'wszystkie'}"
                            cell.value = safe_str(period_text)
                        # POPRAWKA: Usuń skomplikowane formuły SUMPRODUCT które mogą powodować błędy
                        elif header == 'Imie i nazwisko' and len(dataframe) > 0:
                            cell.value = f"Liczba klientow: {len(set(order.customer_name for order in orders if order.customer_name))}"
                        elif header == 'Status' and len(dataframe) > 0:
                            unique_orders_count = len(set(order.baselinker_order_id or f"manual_{order.id}" for order in orders))
                            cell.value = f"Liczba zamowien: {unique_orders_count}"
                            
                except Exception as e:
                    reports_logger.warning(f"Błąd tworzenia formuły dla kolumny {col_idx}: {e}")
                    continue
            
            # WIERSZE 4+: DANE
            for row_idx, row_data in enumerate(dataframe.itertuples(index=False), 4):
                for col_idx, value in enumerate(row_data, 1):
                    try:
                        # POPRAWKA: Bezpieczne wstawianie wartości
                        safe_value = value
                        if isinstance(value, str):
                            safe_value = safe_str(value)
                        elif pd.isna(value):
                            safe_value = ''
                            
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=safe_value)
                        cell.border = border
                        
                        # Kolor tła danych (bardzo jasny)
                        header = headers[col_idx - 1]
                        color_key = COLUMN_COLORS.get(header, 'order_data')
                        light_colors = {
                            'order_data': 'F3F8FF',
                            'customer_data': 'F1F8F1',
                            'logistics_data': 'FFFEF7',
                            'product_data': 'FFF8F0',
                            'financial_data': 'FAF4FB',
                            'production_data': 'FAFAFA'
                        }
                        cell.fill = PatternFill(start_color=light_colors[color_key], end_color=light_colors[color_key], fill_type='solid')
                        
                        # Format liczbowy dla danych
                        if isinstance(value, (int, float)) and value != 0:
                            if 'zl' in header or 'Kwota' in header or 'Wartosc' in header or 'Cena' in header or 'Koszt' in header or 'Zaplacono' in header or 'Saldo' in header:
                                cell.number_format = '#,##0.00" zl"'
                            elif 'm3' in header or 'Objetosc' in header or header in ['Ilosc w produkcji', 'Wyprodukowano', 'Gotowe do odbioru']:  # DODANO "Wyprodukowano"
                                cell.number_format = '#,##0.0000'  # 4 miejsca po przecinku
                            elif header in ['Dlugosc', 'Szerokosc', 'Grubosc']:
                                cell.number_format = '#,##0.00'
                    except Exception as e:
                        reports_logger.warning(f"Błąd wstawiania danych wiersz {row_idx}, kolumna {col_idx}: {e}")
                        continue
            
            # AUTO-DOPASOWANIE SZEROKOŚCI KOLUMN I UKRYWANIE
            for col_idx, header in enumerate(headers, 1):
                try:
                    col_letter = get_column_letter(col_idx)
            
                    # Oblicz maksymalną szerokość na podstawie zawartości
                    max_length = len(str(header))
                    for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=min(worksheet.max_row, 100)):  # Ograniczyć sprawdzanie do 100 wierszy
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
            
                    # Ustaw szerokość (minimum 10, maksimum 30)
                    width = min(max(max_length + 2, 10), 30)
                    worksheet.column_dimensions[col_letter].width = width
            
                    # UKRYWANIE WYBRANYCH KOLUMN (zaktualizowane nazwy bez polskich znaków)
                    columns_to_hide = [
                        'Nr Baselinker',
                        'Nr wew.',
                        'Nazwa klienta',  # POPRAWIONE: było 'Imie i nazwisko'
                        'Kod pocztowy',
                        'Miejscowosc',
                        'Ulica',
                        'Wojewodztwo',
                        'Telefon',
                        'Opiekun',
                        'Dostawa',
                        'Zrodlo',
                        'Grupa',
                        'Rodzaj',
                        'Dlugosc',
                        'Szerokosc',
                        'Ilosc',
                        'Cena brutto',
                        'Cena netto',
                        'Wartosc brutto',
                        'Wartosc netto',
                        'Objetosc 1 szt.',
                        'Objetosc TTL',
                        'Data realizacji',
                        'Status',
                        'Sposob platnosci',
                        'Koszt kuriera',
                        'Zaplacono netto'
                    ]
            
                    # Ukryj kolumnę jeśli jest na liście
                    if header in columns_to_hide:
                        worksheet.column_dimensions[col_letter].hidden = True
                except Exception as e:
                    reports_logger.warning(f"Błąd formatowania kolumny {col_idx}: {e}")
                    continue
            
            # ZAMROŻENIE PANELI (pierwsze 3 wiersze i pierwsze 5 kolumn)
            try:
                worksheet.freeze_panes = 'F4'
            except:
                pass
            
            # FILTRY AUTOMATYCZNE
            try:
                worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{worksheet.max_row}"
            except:
                pass
        
        # UPROSZCZONA FUNKCJA SCALANIA (wykonuj PRZED stylowaniem)
        def add_cell_merging():
            """Uproszczone scalanie komórek - tylko podstawowe"""
            try:
                if len(orders) == 0:
                    return

                # Grupuj dane po zamówieniach
                orders_grouped = {}
                for idx, order in enumerate(orders):
                    order_id = order.baselinker_order_id or f"manual_{order.id}"
                    if order_id not in orders_grouped:
                        orders_grouped[order_id] = []
                    orders_grouped[order_id].append(idx + 4)  # +4 bo dane zaczynają się od wiersza 4
        
                # Uproszczone scalanie - podstawowe kolumny + finansowe
                basic_merge_columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'AI', 'AF', 'AJ', 'AK', 'AL']
        
                for order_id, row_indices in orders_grouped.items():
                    if len(row_indices) > 1:
                        start_row = min(row_indices)
                        end_row = max(row_indices)
                
                        for col_letter in basic_merge_columns:
                            try:
                                merge_range = f'{col_letter}{start_row}:{col_letter}{end_row}'
                                ws_details.merge_cells(merge_range)
                                merged_cell = ws_details[f'{col_letter}{start_row}']
                                # POPRAWKA: Wyśrodkowanie jak reszta komórek
                                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                            except Exception:
                                continue
            except Exception as e:
                reports_logger.warning(f"Błąd scalania komórek: {e}")
                pass

        # WAŻNE: Wywołaj funkcję scalania PRZED formatowaniem
        add_cell_merging()

        # Zastosuj formatowanie do arkusza głównego PO scalaniu
        format_details_sheet(ws_details, df)
        
        # ===== UPROSZCZONY ARKUSZ PODSUMOWANIA =====
        try:
            ws_summary = workbook.create_sheet(title="Podsumowanie")
            
            # Oblicz statystyki
            stats = BaselinkerReportOrder.get_statistics(query)
            
            # Podstawowe statystyki
            unique_customers = len(set(order.customer_name for order in orders if order.customer_name))
            unique_orders = len(set(order.baselinker_order_id or f"manual_{order.id}" for order in orders))
            total_products = len(orders)
            
            # Tytuł
            title_cell = ws_summary.cell(row=1, column=1, value="RAPORT SPRZEDAZY - PODSUMOWANIE")
            title_cell.font = Font(size=16, bold=True, color='1976D2')
            
            # Podstawowe statystyki
            ws_summary.cell(row=3, column=1, value="Liczba zamowien:").font = Font(bold=True)
            ws_summary.cell(row=3, column=2, value=unique_orders)
            
            ws_summary.cell(row=4, column=1, value="Liczba produktow:").font = Font(bold=True)
            ws_summary.cell(row=4, column=2, value=total_products)
            
            ws_summary.cell(row=5, column=1, value="Liczba klientow:").font = Font(bold=True)
            ws_summary.cell(row=5, column=2, value=unique_customers)
            
            ws_summary.cell(row=6, column=1, value="Wartosc netto:").font = Font(bold=True)
            value_cell = ws_summary.cell(row=6, column=2, value=stats['value_net'])
            value_cell.number_format = '#,##0.00" zl"'
            
            ws_summary.cell(row=7, column=1, value="Laczna objetosc:").font = Font(bold=True)
            volume_cell = ws_summary.cell(row=7, column=2, value=stats['total_m3'])
            volume_cell.number_format = '#,##0.0000" m3"'
            
        except Exception as e:
            reports_logger.warning(f"Błąd tworzenia arkusza podsumowania: {e}")
        
        # ===== UPROSZCZONY ARKUSZ KLIENTÓW =====
        try:
            ws_customers = workbook.create_sheet(title="Analiza klientow")

            if len(orders) == 0:
                # Dodaj informację o braku danych
                cell = ws_customers.cell(row=1, column=1, value="Brak danych do analizy klientow")
                cell.font = Font(bold=True, size=14, color='FF0000')
            else:
                # Przygotowanie danych klientów (uproszczone)
                customer_data = {}
                
                for order in orders:
                    customer_name = safe_str(order.customer_name) or 'Nieznany klient'
                    
                    if customer_name not in customer_data:
                        customer_data[customer_name] = {
                            'orders_count': set(),
                            'products_count': 0,
                            'total_value_net': 0,
                            'delivery_state': safe_str(order.delivery_state) or 'Brak danych'
                        }
                    
                    # Aktualizuj dane klienta
                    client = customer_data[customer_name]
                    
                    # Dodaj zamówienie do setu (unikalne ID)
                    if order.baselinker_order_id:
                        client['orders_count'].add(order.baselinker_order_id)
                    else:
                        client['orders_count'].add(f"manual_{order.id}")
                    
                    client['products_count'] += 1
                    client['total_value_net'] += float(order.value_net or 0)
                
                # Konwertuj sety na liczby
                for client_name, data in customer_data.items():
                    data['orders_count'] = len(data['orders_count'])
                
                # Posortuj klientów po wartości
                sorted_customers = sorted(customer_data.items(), key=lambda x: x[1]['total_value_net'], reverse=True)
                
                # Tytuł
                title_cell = ws_customers.cell(row=1, column=1, value="ANALIZA KLIENTOW")
                title_cell.font = Font(size=16, bold=True, color='1976D2')
                
                # Nagłówki tabeli
                headers = ['Lp.', 'Klient', 'Zamowienia', 'Produkty', 'Wartosc netto', 'Wojewodztwo']
                for col, header in enumerate(headers, 1):
                    cell = ws_customers.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
                    cell.border = border
                
                # Dane klientów (TOP 30)
                for rank, (client_name, data) in enumerate(sorted_customers[:30], 1):
                    row_data = [
                        rank,
                        client_name,
                        data['orders_count'],
                        data['products_count'],
                        data['total_value_net'],
                        data['delivery_state']
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws_customers.cell(row=rank + 3, column=col, value=value)
                        cell.border = border
                        
                        # Formatowanie kwot
                        if col == 5:  # Wartość netto
                            cell.number_format = '#,##0.00" zl"'
                
                # AUTO-DOPASOWANIE SZEROKOŚCI KOLUMN
                for col in range(1, len(headers) + 1):
                    col_letter = get_column_letter(col)
                    if col == 2:  # Nazwa klienta
                        width = 25
                    elif col == 6:  # Województwo
                        width = 15
                    else:
                        width = 12
                    ws_customers.column_dimensions[col_letter].width = width
                    
        except Exception as e:
            reports_logger.warning(f"Błąd tworzenia arkusza klientów: {e}")
        
        # ===== BEZPIECZNY ZAPIS DO PAMIĘCI =====
        try:
            workbook.save(output)
            output.seek(0)
        except Exception as e:
            reports_logger.error(f"Błąd zapisywania workbook: {e}")
            return jsonify({
                'success': False,
                'error': f'Błąd zapisywania pliku Excel: {str(e)}'
            }), 500
        
        # Nazwa pliku (bez polskich znaków)
        date_suffix = ""
        if date_from and date_to:
            if date_from == date_to:
                date_suffix = f"_{date_from.strftime('%d-%m-%Y')}"
            else:
                date_suffix = f"_{date_from.strftime('%d-%m-%Y')}_{date_to.strftime('%d-%m-%Y')}"
        
        filename = f"raporty_sprzedazy{date_suffix}_{datetime.now().strftime('%d-%m-%Y_%H%M%S')}.xlsx"
        
        reports_logger.info("Wygenerowano ulepszony eksport Excel",
                          user_email=user_email,
                          records_count=len(excel_data),
                          filename=filename)
        
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )
        
    except Exception as e:
        reports_logger.error("Błąd eksportu Excel",
                           user_email=user_email,
                           error=str(e))
        return jsonify({
            'success': False,
            'error': f'Błąd eksportu: {str(e)}'
        }), 500

@reports_bp.route('/api/dropdown-values/<field_name>')
@login_required
def api_get_dropdown_values(field_name):
    """
    API endpoint do pobierania unikalnych wartości dla dropdown'ów
    """
    try:
        if not hasattr(BaselinkerReportOrder, field_name):
            return jsonify({'success': False, 'error': 'Nieprawidłowe pole'}), 400
        
        # Pobierz unikalne wartości z bazy (wykluczając anulowane i nieopłacone)
        excluded_statuses = ['Zamówienie anulowane', 'Nowe - nieopłacone']
        column = getattr(BaselinkerReportOrder, field_name)
        
        query = db.session.query(column)\
            .filter(column.isnot(None))\
            .filter(~BaselinkerReportOrder.current_status.in_(excluded_statuses))\
            .distinct()
        
        values = query.all()
        
        # Wyciągnij wartości z tupli
        unique_values = sorted([value[0] for value in values if value[0]])
        
        return jsonify({
            'success': True,
            'values': unique_values
        })
        
    except Exception as e:
        reports_logger.error("Błąd pobierania dropdown values",
                           field_name=field_name,
                           error=str(e))
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ===== FUNKCJE POMOCNICZE =====

def _sync_selected_orders(service: BaselinkerReportsService, order_ids: List[int]) -> Dict:
    """
    Synchronizuje wybrane zamówienia - pobiera pełne informacje i aktualizuje wszystkie dane
    
    Args:
        service: Serwis Baselinker
        order_ids: Lista ID zamówień do synchronizacji
        
    Returns:
        Dict: Wynik synchronizacji
    """
    try:
        reports_logger.info("Rozpoczęcie synchronizacji wybranych zamówień",
                          order_ids_count=len(order_ids),
                          order_ids=order_ids[:10])  # Loguj pierwsze 10
        
        orders = []
        failed_orders = []
        
        # Pobierz pełne dane każdego zamówienia
        for order_id in order_ids:
            try:
                order = service.get_order_details(order_id)
                if order:
                    orders.append(order)
                    reports_logger.debug("Pobrano szczegóły zamówienia",
                                       order_id=order_id,
                                       products_count=len(order.get('products', [])))
                else:
                    failed_orders.append(order_id)
                    reports_logger.warning("Nie udało się pobrać zamówienia",
                                         order_id=order_id)
            except Exception as e:
                failed_orders.append(order_id)
                reports_logger.error("Błąd pobierania zamówienia",
                                   order_id=order_id,
                                   error=str(e))
        
        if not orders:
            error_msg = f'Nie udało się pobrać żadnego z wybranych zamówień. Nieudane: {failed_orders}'
            reports_logger.error("Brak pobranych zamówień", 
                               failed_orders=failed_orders,
                               total_requested=len(order_ids))
            return {
                'success': False,
                'error': error_msg
            }
        
        # ZMIANA: Używamy pełnej synchronizacji zamiast tylko dodawania
        # To oznacza że aktualizowane są wszystkie informacje w istniejących rekordach
        reports_logger.info("Rozpoczęcie pełnej synchronizacji zamówień",
                          orders_to_sync=len(orders),
                          failed_orders_count=len(failed_orders))
        
        # Użyj metody sync_orders która obsługuje aktualizacje
        result = service.sync_orders(orders_list=orders, sync_type='selected')
        
        # Dodaj informacje o nieudanych zamówieniach do wyniku
        result['failed_orders'] = failed_orders
        result['failed_orders_count'] = len(failed_orders)
        
        if failed_orders:
            success_count = result.get('orders_processed', 0)
            total_requested = len(order_ids)
            result['message'] = f'Zsynchronizowano {success_count} z {total_requested} zamówień. Nieudane: {len(failed_orders)}'
            
            reports_logger.warning("Synchronizacja częściowo nieudana",
                                 success_count=success_count,
                                 total_requested=total_requested,
                                 failed_count=len(failed_orders),
                                 failed_orders=failed_orders)
        else:
            reports_logger.info("Synchronizacja wybranych zamówień zakończona pomyślnie",
                              orders_processed=result.get('orders_processed', 0),
                              orders_added=result.get('orders_added', 0),
                              orders_updated=result.get('orders_updated', 0))
        
        return result
            
    except Exception as e:
        reports_logger.error("Błąd synchronizacji wybranych zamówień",
                           error=str(e),
                           error_type=type(e).__name__,
                           order_ids_count=len(order_ids))
        return {
            'success': False,
            'error': str(e)
        }

# Synchronizacja statusów
@reports_bp.route('/api/sync-statuses', methods=['POST'])
@login_required
def api_sync_statuses():
    """
    API endpoint do synchronizacji statusów zamówień z Baselinker
    """
    user_email = session.get('user_email')
    
    try:
        reports_logger.info("Rozpoczęcie synchronizacji statusów",
                          user_email=user_email)
        
        # Pobierz serwis
        service = get_reports_service()
        
        # ZMIANA: Pobierz zamówienia które mogą być synchronizowane (wykluczamy tylko 105112 i 138625)
        # 105112 = "Nowe - nieopłacone"
        # 138625 = "Zamówienie anulowane"
        excluded_status_ids = [105112, 138625]
        excluded_status_names = [
            'Nowe - nieopłacone',
            'Zamówienie anulowane'
        ]
        
        # Pobierz zamówienia z bazy które nie mają wykluczonych statusów
        orders_to_sync = BaselinkerReportOrder.query.filter(
            BaselinkerReportOrder.baselinker_order_id.isnot(None),
            ~BaselinkerReportOrder.baselinker_status_id.in_(excluded_status_ids),
            ~BaselinkerReportOrder.current_status.in_(excluded_status_names)
        ).all()
        
        if not orders_to_sync:
            reports_logger.info("Brak zamówień do synchronizacji statusów",
                              user_email=user_email,
                              excluded_status_ids=excluded_status_ids,
                              excluded_status_names=excluded_status_names)
            return jsonify({
                'success': True,
                'message': 'Brak zamówień do synchronizacji statusów',
                'orders_processed': 0,
                'orders_updated': 0
            })
        
        # Grupuj zamówienia według baselinker_order_id
        unique_order_ids = list(set(order.baselinker_order_id for order in orders_to_sync))
        
        reports_logger.info("Synchronizacja statusów zamówień",
                          user_email=user_email,
                          unique_orders=len(unique_order_ids),
                          total_records=len(orders_to_sync),
                          excluded_status_ids=excluded_status_ids)
        
        # Synchronizuj statusy
        updated_count = 0
        processed_count = 0
        payment_updated_count = 0
        status_updated_count = 0
        internal_number_updated_count = 0
        delivery_updated_count = 0
        sync_start = datetime.utcnow()
        
        for order_id in unique_order_ids:
            try:
                # Pobierz aktualny status z Baselinker
                order_details = service.get_order_details(order_id)
                
                if order_details:
                    # Pobierz nowy status
                    new_status_id = order_details.get('order_status_id')
                    new_status = service.status_map.get(new_status_id, f'Status {new_status_id}')
                    
                    # Pobierz kwotę zapłaconą (brutto -> netto)
                    payment_done = order_details.get('payment_done', 0)
                    custom_fields = order_details.get('custom_extra_fields', {})
                    price_type_from_api = custom_fields.get('106169', '').strip()
                    
                    new_paid_amount_net = service._calculate_paid_amount_net(payment_done, price_type_from_api)
                    
                    # NOWE: Pobierz numer wewnętrzny z extra_field_1
                    new_internal_number = order_details.get('extra_field_1', '').strip()
                    
                    # NOWE: Pobierz dane dostawy
                    new_delivery_method = order_details.get('delivery_method', '').strip()
                    new_delivery_cost_gross = float(order_details.get('delivery_price', 0))
                    
                    records = BaselinkerReportOrder.query.filter_by(
                        baselinker_order_id=order_id
                    ).all()

                    records_updated = 0
                    for record in records:
                        record.current_status = new_status
                        record.baselinker_status_id = new_status_id
                        record.paid_amount_net = new_paid_amount_net
                        record.internal_order_number = new_internal_number
                        record.delivery_method = new_delivery_method
                        record.delivery_cost = new_delivery_cost_gross
                        record.updated_at = datetime.utcnow()
                        record.update_production_fields()
                        records_updated += 1

                    if records_updated > 0:
                        updated_count += records_updated
                        status_updated_count += 1

                        if new_paid_amount_net > 0:
                            payment_updated_count += 1

                        if new_internal_number:
                            internal_number_updated_count += 1

                        if new_delivery_method or new_delivery_cost_gross > 0:
                            delivery_updated_count += 1

                        reports_logger.debug("Zaktualizowano zamówienie",
                                           order_id=order_id,
                                           new_status=new_status,
                                           new_status_id=new_status_id,
                                           paid_amount_net=new_paid_amount_net,
                                           internal_number=new_internal_number,
                                           delivery_method=new_delivery_method,
                                           delivery_cost=new_delivery_cost_gross,
                                           records_updated=records_updated)
                    
                    processed_count += 1
                else:
                    reports_logger.warning("Nie udało się pobrać szczegółów zamówienia",
                                         order_id=order_id)
                    
            except Exception as e:
                reports_logger.error("Błąd synchronizacji statusu zamówienia",
                                   order_id=order_id,
                                   error=str(e))
                continue
        
        # Zapisz zmiany
        db.session.commit()
        
        duration = (datetime.utcnow() - sync_start).total_seconds()
        
        reports_logger.info("Synchronizacja statusów zakończona",
                          user_email=user_email,
                          processed_orders=processed_count,
                          updated_records=updated_count,
                          status_updated_count=status_updated_count,
                          payment_updated_count=payment_updated_count,
                          internal_number_updated_count=internal_number_updated_count,
                          delivery_updated_count=delivery_updated_count,
                          duration_seconds=duration)
        
        return jsonify({
            'success': True,
            'message': f'Zsynchronizowano statusy {processed_count} zamówień',
            'orders_processed': processed_count,
            'orders_updated': status_updated_count,
            'records_updated': updated_count,
            'payment_updated_count': payment_updated_count,
            'internal_number_updated': internal_number_updated_count,
            'delivery_updated': delivery_updated_count
        })
        
    except Exception as e:
        db.session.rollback()
        reports_logger.error("Błąd synchronizacji statusów",
                           user_email=user_email,
                           error=str(e),
                           error_type=type(e).__name__)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@reports_bp.route('/api/delete-manual-row', methods=['POST'])
@login_required
def api_delete_manual_row():
    """
    API endpoint do usuwania rekordów z bazy danych
    Obsługuje usuwanie pojedynczych rekordów oraz całych zamówień wieloproduktowych
    """
    user_email = session.get('user_email')
    
    try:
        data = request.get_json()
        record_id = data.get('record_id')
        delete_all_products = data.get('delete_all_products', False)
        
        if not record_id:
            return jsonify({'success': False, 'error': 'Brak ID rekordu'}), 400
        
        reports_logger.info("Rozpoczęcie usuwania rekordu",
                          user_email=user_email,
                          record_id=record_id,
                          delete_all_products=delete_all_products)
        
        # Pobierz główny rekord
        main_record = BaselinkerReportOrder.query.get_or_404(record_id)
        
        records_to_delete = []
        
        if delete_all_products and main_record.baselinker_order_id:
            # Usuń wszystkie produkty tego zamówienia z Baselinker
            all_order_records = BaselinkerReportOrder.query.filter_by(
                baselinker_order_id=main_record.baselinker_order_id
            ).all()
            
            records_to_delete = all_order_records
            
            reports_logger.info("Usuwanie całego zamówienia wieloproduktowego",
                              user_email=user_email,
                              baselinker_order_id=main_record.baselinker_order_id,
                              products_count=len(all_order_records))
        else:
            # Usuń tylko pojedynczy rekord
            records_to_delete = [main_record]
            
            reports_logger.info("Usuwanie pojedynczego rekordu",
                              user_email=user_email,
                              record_id=record_id,
                              is_manual=main_record.is_manual)
        
        # Zbierz informacje o usuwanych rekordach (dla logowania)
        deleted_info = []
        for record in records_to_delete:
            deleted_info.append({
                'id': record.id,
                'customer_name': record.customer_name,
                'is_manual': record.is_manual,
                'baselinker_order_id': record.baselinker_order_id,
                'date_created': record.date_created.isoformat() if record.date_created else None
            })
        
        # Usuń rekordy z bazy danych
        deleted_count = 0
        for record in records_to_delete:
            db.session.delete(record)
            deleted_count += 1
        
        # Zatwierdź transakcję
        db.session.commit()
        
        reports_logger.info("Pomyślnie usunięto rekordy",
                          user_email=user_email,
                          deleted_count=deleted_count,
                          deleted_records=deleted_info)
        
        # Przygotuj komunikat zwrotny
        if deleted_count > 1:
            message = f'Usunięto zamówienie z {deleted_count} produktami'
        else:
            record_type = "ręczny rekord" if main_record.is_manual else "rekord z Baselinker"
            message = f'Usunięto {record_type}'
        
        return jsonify({
            'success': True,
            'message': message,
            'deleted_count': deleted_count,
            'deleted_records': [info['id'] for info in deleted_info]
        })
        
    except Exception as e:
        db.session.rollback()
        reports_logger.error("Błąd usuwania rekordu",
                           user_email=user_email,
                           record_id=record_id,
                           error=str(e),
                           error_type=type(e).__name__)
        return jsonify({
            'success': False,
            'error': f'Błąd usuwania rekordu: {str(e)}'
        }), 500
    
@reports_bp.route('/api/fetch-orders-for-selection', methods=['POST'])
@login_required
def api_fetch_orders_for_selection():
    """
    POPRAWIONY ENDPOINT: Fetches orders from Baselinker for selected date range
    with automatic pagination when >90 orders
    DEFAULT EXCLUDES STATUSES 105112 and 138625
    + Volume and attributes analysis for products
    + NOWE: Całkowicie ignoruje zamówienia które już istnieją w bazie danych
    """
    user_email = session.get('user_email')
    
    try:
        data = request.get_json()
        if not data:
            reports_logger.error("Missing data in fetch-orders-for-selection request", 
                               user_email=user_email)
            return jsonify({
                'success': False,
                'error': 'Missing data in request'
            }), 400

        date_from = data.get('date_from')
        date_to = data.get('date_to')
        days_count = data.get('days_count')
        get_all_statuses = data.get('get_all_statuses', False)

        if not all([date_from, date_to, days_count]):
            return jsonify({
                'success': False,
                'error': 'Missing required parameters: date_from, date_to, days_count'
            }), 400

        reports_logger.info("Fetching orders for selection with volume analysis",
                          user_email=user_email,
                          date_from=date_from,
                          date_to=date_to,
                          days_count=days_count,
                          get_all_statuses=get_all_statuses)

        service = get_reports_service()
        
        # Get existing orders from database to avoid duplicates
        existing_orders = BaselinkerReportOrder.query.filter(
            BaselinkerReportOrder.baselinker_order_id.isnot(None)
        ).with_entities(BaselinkerReportOrder.baselinker_order_id).distinct().all()
        existing_order_ids = {order[0] for order in existing_orders}
        
        reports_logger.info("Loaded existing orders from database", 
                         existing_count=len(existing_order_ids))
        
        # NOWA LOGIKA: Pobierz wszystkie zamówienia w jednym wywołaniu z poprawną paginacją
        reports_logger.info("Fetching all orders from date range using corrected pagination",
                          date_from=date_from,
                          date_to=date_to)
        
        # Konwertuj daty
        start_date = datetime.fromisoformat(date_from)
        end_date = datetime.fromisoformat(date_to)
        
        # KLUCZ: Użyj fetch_orders_from_date_range zamiast wielokrotnego wywołania fetch_orders_from_baselinker
        result = service.fetch_orders_from_date_range(
            date_from=start_date,
            date_to=end_date,
            get_all_statuses=get_all_statuses
        )
        
        if not result['success']:
            return jsonify({
                'success': False,
                'error': result.get('error', 'Unknown error fetching orders')
            }), 500
        
        all_orders = result['orders']
        reports_logger.info(f"Fetched {len(all_orders)} orders from date range")

        # Dodatkowe filtrowanie po dacie złożenia zamówienia (date_add)
        start_ts = int(start_date.timestamp())
        end_ts = int(end_date.timestamp()) + 86399
        filtered_orders = []
        for order in all_orders:
            order_date_add = order.get('date_add')
            try:
                order_date_add = int(order_date_add)
            except (TypeError, ValueError):
                continue
            if start_ts <= order_date_add <= end_ts:
                filtered_orders.append(order)

        reports_logger.info("Orders after date_add filtering",
                             before=len(all_orders), after=len(filtered_orders))
        all_orders = filtered_orders
        
        # ZMIANA: Filtruj zamówienia - pokaż TYLKO te które NIE istnieją w bazie danych
        processed_orders = []
        new_orders_count = 0
        ignored_existing_count = 0
        
        for order in all_orders:
            order_id = order['order_id']
            
            # KLUCZOWA ZMIANA: Sprawdź czy zamówienie już istnieje - jeśli TAK, IGNORUJ całkowicie
            if order_id in existing_order_ids:
                ignored_existing_count += 1
                reports_logger.debug("Ignoring existing order",
                                   order_id=order_id,
                                   customer_name=order.get('delivery_fullname', 'Brak nazwy'))
                continue  # Pomiń to zamówienie całkowicie
                
            # Dodatkowa ochrona: filtruj wykluczone statusy po stronie aplikacji
            status_id = order.get('order_status_id')
            if not get_all_statuses and status_id in [105112, 138625]:
                reports_logger.debug("Excluded order due to status",
                                   order_id=order_id,
                                   status_id=status_id,
                                   status_name=service.status_map.get(status_id, f'Status {status_id}'))
                continue
            
            # Wykonaj analizę objętości dla produktów w zamówieniu
            order = analyze_order_products_for_volume(order)
            
            # Oznacz jako nowe zamówienie (wszystkie tutaj są nowe, bo istniejące zostały pominięte)
            order['exists_in_database'] = False
            
            processed_orders.append(order)
            new_orders_count += 1

        # Analiza objętości
        total_volume_issues = sum(1 for order in processed_orders if order.get('has_volume_issues', False))

        reports_logger.info("Completed fetching orders with volume analysis",
                          total_orders_fetched=len(all_orders),
                          ignored_existing=ignored_existing_count,
                          new_orders_displayed=new_orders_count,
                          volume_issues_count=total_volume_issues)

        # Sprawdź czy są jakieś nowe zamówienia do wyświetlenia
        if new_orders_count == 0:
            return jsonify({
                'success': True,
                'orders': [],
                'total_orders': 0,
                'new_orders': 0,
                'ignored_existing': ignored_existing_count,
                'volume_issues_count': 0,
                'message': f'Brak nowych zamówień w wybranym okresie. Zignorowano {ignored_existing_count} zamówień już istniejących w bazie danych.'
            })

        return jsonify({
            'success': True,
            'orders': processed_orders,
            'total_orders': len(processed_orders),
            'new_orders': new_orders_count,
            'ignored_existing': ignored_existing_count,
            'volume_issues_count': total_volume_issues,
            'pagination_info': {
                'method': 'date_range_fetch',
                'filtered_excluded_statuses': not get_all_statuses
            },
            'message': f'Znaleziono {new_orders_count} nowych zamówień. Zignorowano {ignored_existing_count} już istniejących. {total_volume_issues} produktów wymaga uzupełnienia objętości.'
        })

    except Exception as e:
        reports_logger.error("Error in fetch-orders-for-selection",
                           user_email=user_email,
                           error=str(e),
                           error_type=type(e).__name__)
        return jsonify({
            'success': False,
            'error': f'Błąd pobierania zamówień: {str(e)}'
        }), 500

def analyze_order_products_for_volume(order_data):
    """
    ROZSZERZONA FUNKCJA: Analizuje produkty w zamówieniu pod kątem objętości i atrybutów
    DODANO: Obsługę worków opałowych (bez walidacji)
    """
    products = order_data.get('products', [])
    if not products:
        order_data['has_volume_issues'] = False
        return order_data
    
    order_has_volume_issues = False
    analyzed_products = []
    
    for product in products:
        product_name = product.get('name', '')
    
        # Sprawdź czy to usługa używając service
        service = get_reports_service()
        if service._is_service_product(product_name):
            analysis = {
                'analysis_type': 'service',
                'has_dimensions': False,
                'has_volume': False, 
                'volume': None,
                'wood_species': None,
                'technology': None,
                'wood_class': None
            }
            # Usługi nie wymagają uzupełnienia objętości
            product['needs_manual_volume'] = False
            product['has_dimension_issues'] = False
        else:
            # Analizuj produkty fizyczne
            analysis = analyze_product_for_volume_and_attributes(product_name)
            
            # NOWA LOGIKA: Worki opałowe nie wymagają walidacji
            if analysis['analysis_type'] == 'no_validation_needed':
                product['needs_manual_volume'] = False
                product['has_dimension_issues'] = False
            else:
                product['needs_manual_volume'] = analysis['analysis_type'] == 'manual_input_needed'
                # Sprawdź czy trzeba też sprawdzić wymiary (stara logika)
                product['has_dimension_issues'] = not check_product_dimensions(product_name)
    
        # Dodaj wyniki analizy do produktu
        product['volume_analysis'] = analysis
    
        # ZMIANA: Worki opałowe nie powodują problemów z objętością
        if analysis['analysis_type'] == 'manual_input_needed' or analysis['analysis_type'] == 'volume_only':
            order_has_volume_issues = True
    
        analyzed_products.append(product)
    
    # Aktualizuj zamówienie
    order_data['products'] = analyzed_products
    order_data['has_volume_issues'] = order_has_volume_issues
    
    # Zachowaj istniejącą logikę dla has_dimension_issues
    order_data['has_dimension_issues'] = any(p.get('has_dimension_issues', False) for p in analyzed_products)
    
    return order_data
    
def _sync_selected_orders_with_volumes(service, order_ids):
    """
    NOWA FUNKCJA: Synchronizuje wybrane zamówienia z obsługą objętości i atrybutów
    """
    try:
        orders_processed = 0
        orders_added = 0
        
        for order_id in order_ids:
            print(f"[DEBUG] Przetwarzanie zamówienia {order_id} z objętościami")
            
            # Pobierz zamówienie z Baselinker
            order_data = service.get_single_order_from_baselinker(order_id)
            
            if not order_data:
                print(f"[WARNING] Nie można pobrać zamówienia {order_id}")
                continue
            
            # Przetwórz produkty w zamówieniu
            for product in order_data.get('products', []):
                # Przeprowadź analizę produktu
                analysis = analyze_product_for_volume_and_attributes(product.get('name', ''))
                
                # Przygotuj dane do zapisania
                record_data = service.prepare_order_record_data(order_data, product)
                
                # NOWA LOGIKA: Obsługa objętości i atrybutów
                if analysis['analysis_type'] == 'dimensions_priority':
                    # Wymiary mają priorytet - oblicz objętość standardowo
                    volume = service.calculate_volume_from_dimensions(
                        record_data.get('length_cm', 0),
                        record_data.get('width_cm', 0), 
                        record_data.get('thickness_cm', 0),
                        record_data.get('quantity', 1)
                    )
                    record_data['total_volume'] = volume
                    
                elif analysis['analysis_type'] == 'volume_only':
                    # ✅ POPRAWKA: objętość z nazwy to już total_volume całej pozycji
                    total_volume = float(analysis.get('volume', 0))
                    quantity = int(record_data.get('quantity', 1))
    
                    record_data['total_volume'] = round(total_volume, 4)  # NIE MNÓŻ!
                    record_data['volume_per_piece'] = round(total_volume / quantity, 4)  # PODZIEL!
                    
                    # Wyczyść wymiary (bo ich nie ma)
                    record_data['length_cm'] = None
                    record_data['width_cm'] = None
                    record_data['thickness_cm'] = None
                    
                    print(f"[DEBUG] volume_only: {total_volume} m³ (całość) / {quantity} = {record_data['volume_per_piece']} m³/szt")
                    
                elif analysis['analysis_type'] == 'manual_input_needed':
                    # Użyj ręcznie wprowadzonych danych
                    product_key = f"{order_id}_{product.get('product_id', 'unknown')}"
                    volume_fix = service.get_volume_fix(product_key)
                    
                    if volume_fix and 'volume' in volume_fix:
                        volume_per_piece = float(volume_fix['volume'])
                        quantity = record_data.get('quantity', 1)
                        total_volume = float(volume_fix['volume'])

                        record_data['total_volume'] = total_volume  # NIE MNÓŻ przez quantity!
                        record_data['volume_per_piece'] = total_volume / quantity  # Podziel przez quantity
                        
                        # Wyczyść wymiary
                        record_data['length_cm'] = None
                        record_data['width_cm'] = None
                        record_data['thickness_cm'] = None
                    else:
                        # Brak danych - ustaw objętość na 0
                        record_data['total_volume'] = 0
                        record_data['volume_per_piece'] = 0
                
                # Dodaj atrybuty z analizy lub z ręcznego wprowadzenia
                record_data['wood_species'] = analysis.get('wood_species') or service.get_volume_fix_attribute(
                    f"{order_id}_{product.get('product_id', 'unknown')}", 'wood_species'
                )
                record_data['technology'] = analysis.get('technology') or service.get_volume_fix_attribute(
                    f"{order_id}_{product.get('product_id', 'unknown')}", 'technology'
                )
                record_data['wood_class'] = analysis.get('wood_class') or service.get_volume_fix_attribute(
                    f"{order_id}_{product.get('product_id', 'unknown')}", 'wood_class'
                )
                
                # Zapisz rekord do bazy
                service.save_order_record(record_data)
                orders_added += 1
            
            orders_processed += 1
        
        return {
            'success': True,
            'orders_processed': orders_processed,
            'orders_added': orders_added,
            'message': f'Pomyślnie zapisano {orders_added} rekordów z {orders_processed} zamówień'
        }
        
    except Exception as e:
        print(f"[ERROR] Błąd w _sync_selected_orders_with_volumes: {str(e)}")
        return {
            'success': False,
            'error': f'Błąd synchronizacji: {str(e)}'
        }

@reports_bp.route('/api/save-selected-orders-with-dimensions', methods=['POST'])
@login_required
def api_save_selected_orders_with_dimensions():
    """
    NOWY ENDPOINT: Zapisuje wybrane zamówienia z opcjonalnym uzupełnieniem wymiarów
    """
    user_email = session.get('user_email')
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': 'Brak danych w zapytaniu'
            }), 400

        order_ids = data.get('order_ids', [])
        dimension_fixes = data.get('dimension_fixes', {})  # {order_id: {product_id: {length_cm: X, width_cm: Y, thickness_mm: Z}}}

        if not order_ids:
            return jsonify({
                'success': False,
                'error': 'Brak wybranych zamówień do zapisania'
            }), 400

        reports_logger.info("Rozpoczęcie zapisywania zamówień z wymiarami",
                          user_email=user_email,
                          order_ids_count=len(order_ids),
                          has_dimension_fixes=bool(dimension_fixes))

        # Walidacja ID zamówień
        try:
            order_ids = [int(order_id) for order_id in order_ids]
        except (ValueError, TypeError) as e:
            return jsonify({
                'success': False,
                'error': 'Błędne ID zamówień'
            }), 400

        # Sprawdź które zamówienia już istnieją w bazie
        existing_orders = BaselinkerReportOrder.query.filter(
            BaselinkerReportOrder.baselinker_order_id.in_(order_ids)
        ).with_entities(BaselinkerReportOrder.baselinker_order_id).distinct().all()
        existing_order_ids = {order.baselinker_order_id for order in existing_orders}

        # Filtruj tylko nowe zamówienia
        new_order_ids = [order_id for order_id in order_ids if order_id not in existing_order_ids]

        if not new_order_ids:
            return jsonify({
                'success': True,
                'message': 'Wszystkie wybrane zamówienia już istnieją w bazie danych',
                'orders_saved': 0,
                'orders_skipped': len(order_ids)
            })

        # Pobierz service
        service = get_reports_service()
        
        # Zastosuj poprawki wymiarów jeśli zostały podane
        if dimension_fixes:
            service.set_dimension_fixes(dimension_fixes)
            reports_logger.info("Zastosowano poprawki wymiarów", 
                              fixes_count=len(dimension_fixes))

        # Synchronizuj wybrane zamówienia
        result = _sync_selected_orders(service, new_order_ids)
        
        # Wyczyść poprawki wymiarów
        if dimension_fixes:
            service.clear_dimension_fixes()

        if result.get('success'):
            reports_logger.info("Zapisywanie zamówień z wymiarami zakończone pomyślnie",
                              orders_processed=result.get('orders_processed', 0),
                              orders_added=result.get('orders_added', 0))
            return jsonify(result)
        else:
            return jsonify(result), 500
            
    except Exception as e:
        reports_logger.error("Błąd zapisywania zamówień z wymiarami",
                           user_email=user_email,
                           error=str(e))
        return jsonify({
            'success': False,
            'error': f'Błąd zapisywania zamówień: {str(e)}'
        }), 500

def check_product_dimensions(product_name):
    """
    NOWA FUNKCJA: Sprawdza czy nazwa produktu zawiera wymiary
    NAPRAWIONE: Obsługuje liczby z przecinkami i kropkami
    
    Args:
        product_name (str): Nazwa produktu
        
    Returns:
        bool: True jeśli produkt ma wymiary, False jeśli nie
    """
    if not product_name:
        return False
    
    name_lower = product_name.lower()
    
    # NAPRAWIONE: Wzorce obsługujące liczby z przecinkami i kropkami
    dimension_patterns = [
        # Klasyczne formaty 3 wymiarów (z kropkami i przecinkami)
        r'\d+[,.]?\d*\s*x\s*\d+[,.]?\d*\s*x\s*\d+[,.]?\d*',  # 200,4x89x4.5
        r'\d+[,.]?\d*\s*×\s*\d+[,.]?\d*\s*×\s*\d+[,.]?\d*',  # 200,4×89×4.5
        
        # Format 2 wymiarów (z kropkami i przecinkami)
        r'\d+[,.]?\d*\s*x\s*\d+[,.]?\d*(?!\s*x)',  # 200,4x89 (ale nie x coś dalej)
        r'\d+[,.]?\d*\s*×\s*\d+[,.]?\d*(?!\s*×)',  # 200,4×89
        
        # Z jednostkami cm/mm
        r'\d+[,.]?\d*\s*cm\s*x\s*\d+[,.]?\d*\s*cm',  # 200,4cm x 89cm
        r'\d+[,.]?\d*\s*mm\s*x\s*\d+[,.]?\d*\s*mm',  # 200,4mm x 89mm
        r'\d+[,.]?\d*\s*cm\s*×\s*\d+[,.]?\d*\s*cm',  # 200,4cm × 89cm
        
        # Wymiary w tekście
        r'długość\s*:?\s*\d+[,.]?\d*',  # "długość: 200,4"
        r'szerokość\s*:?\s*\d+[,.]?\d*',  # "szerokość: 89"
        r'grubość\s*:?\s*\d+[,.]?\d*',  # "grubość: 4,5"
        r'wysokość\s*:?\s*\d+[,.]?\d*',  # "wysokość: 4,5"
        r'głębokość\s*:?\s*\d+[,.]?\d*',  # "głębokość: 4,5"
        
        # Dodatkowe formaty
        r'\d+[,.]?\d*\s*/\s*\d+[,.]?\d*\s*/\s*\d+[,.]?\d*',  # 200,4/89/4.5
        r'\d+[,.]?\d*-\d+[,.]?\d*-\d+[,.]?\d*',  # 200,4-89-4.5
        
        # Wymiary w nawiasach
        r'\(\s*\d+[,.]?\d*\s*x\s*\d+[,.]?\d*.*?\)',  # (200,4 x 89)
        r'\[\s*\d+[,.]?\d*\s*x\s*\d+[,.]?\d*.*?\]',  # [200,4 x 89]
    ]
    
    import re
    for pattern in dimension_patterns:
        if re.search(pattern, name_lower):
            return True
    
    # Dodatkowe sprawdzenia heurystyczne
    has_numbers = any(char.isdigit() for char in product_name)
    has_dimension_separators = any(sep in name_lower for sep in ['x', '×', '/', '-'])
    has_units = any(unit in name_lower for unit in ['cm', 'mm', 'm'])
    
    # Jeśli ma liczby i separatory wymiarów lub jednostki
    if has_numbers and (has_dimension_separators or has_units):
        # Dodatkowa walidacja - sprawdź czy to nie jest tylko data lub numer zamówienia
        date_patterns = [
            r'\d{1,2}[.-/]\d{1,2}[.-/]\d{2,4}',  # Daty
            r'\d{4,}',  # Długie numery (prawdopodobnie ID)
        ]
        
        is_probably_date_or_id = any(re.search(pattern, product_name) for pattern in date_patterns)
        
        if not is_probably_date_or_id:
            return True
    
    print(f"[DEBUG] Brak wymiarów w '{product_name}'")
    return False

@reports_bp.route('/api/save-selected-orders', methods=['POST'])
@login_required
def api_save_selected_orders():
    """
    NOWY ENDPOINT: Zapisuje wybrane zamówienia do bazy danych
    """
    user_email = session.get('user_email')
    
    try:
        data = request.get_json()
        if not data:
            reports_logger.error("Brak danych w zapytaniu save-selected-orders",
                               user_email=user_email)
            return jsonify({
                'success': False,
                'error': 'Brak danych w zapytaniu'
            }), 400

        order_ids = data.get('order_ids', [])
        date_from = data.get('date_from')
        date_to = data.get('date_to')

        if not order_ids:
            return jsonify({
                'success': False,
                'error': 'Brak wybranych zamówień do zapisania'
            }), 400

        reports_logger.info("Rozpoczęcie zapisywania wybranych zamówień",
                          user_email=user_email,
                          order_ids_count=len(order_ids),
                          order_ids=order_ids,
                          date_from=date_from,
                          date_to=date_to)

        # Walidacja ID zamówień
        try:
            order_ids = [int(order_id) for order_id in order_ids]
        except (ValueError, TypeError) as e:
            reports_logger.error("Błędne ID zamówień",
                               user_email=user_email,
                               order_ids=order_ids,
                               error=str(e))
            return jsonify({
                'success': False,
                'error': 'Błędne ID zamówień'
            }), 400

        # Sprawdź które zamówienia już istnieją w bazie
        existing_orders = BaselinkerReportOrder.query.filter(
            BaselinkerReportOrder.baselinker_order_id.in_(order_ids)
        ).with_entities(BaselinkerReportOrder.baselinker_order_id).distinct().all()
        existing_order_ids = {order.baselinker_order_id for order in existing_orders}

        # Filtruj tylko nowe zamówienia
        new_order_ids = [order_id for order_id in order_ids if order_id not in existing_order_ids]

        reports_logger.info("Analiza zamówień do zapisania",
                          total_requested=len(order_ids),
                          existing_in_db=len(existing_order_ids),
                          new_to_save=len(new_order_ids),
                          existing_order_ids=list(existing_order_ids),
                          new_order_ids=new_order_ids)

        if not new_order_ids:
            reports_logger.warning("Wszystkie wybrane zamówienia już istnieją w bazie",
                                 user_email=user_email,
                                 requested_orders=order_ids)
            return jsonify({
                'success': True,
                'message': 'Wszystkie wybrane zamówienia już istnieją w bazie danych',
                'orders_saved': 0,
                'orders_skipped': len(order_ids),
                'existing_orders': order_ids
            })

        # Użyj istniejącej funkcji synchronizacji dla wybranych zamówień
        service = get_reports_service()
        
        reports_logger.info("Wywołanie synchronizacji dla wybranych zamówień",
                          new_order_ids=new_order_ids)

        # Użyj funkcji _sync_selected_orders (może trzeba będzie ją zmodyfikować)
        result = _sync_selected_orders(service, new_order_ids)

        if result.get('success'):
            reports_logger.info("Zapisywanie wybranych zamówień zakończone pomyślnie",
                              user_email=user_email,
                              orders_requested=len(order_ids),
                              orders_saved=result.get('orders_added', 0),
                              orders_updated=result.get('orders_updated', 0),
                              orders_processed=result.get('orders_processed', 0))

            return jsonify({
                'success': True,
                'message': f'Pomyślnie zapisano {result.get("orders_added", 0)} zamówień',
                'orders_saved': result.get('orders_added', 0),
                'orders_updated': result.get('orders_updated', 0),
                'orders_processed': result.get('orders_processed', 0),
                'orders_skipped': len(existing_order_ids),
                'existing_orders': list(existing_order_ids)
            })
        else:
            error_msg = result.get('error', 'Nieznany błąd synchronizacji')
            reports_logger.error("Błąd podczas zapisywania wybranych zamówień",
                               user_email=user_email,
                               error=error_msg)
            return jsonify({
                'success': False,
                'error': f'Błąd zapisywania: {error_msg}'
            }), 500

    except Exception as e:
        reports_logger.error("Błąd zapisywania wybranych zamówień",
                           user_email=user_email,
                           error=str(e),
                           error_type=type(e).__name__)
        return jsonify({
            'success': False,
            'error': f'Błąd serwera: {str(e)}'
        }), 500

@reports_bp.route('/api/export-routimo', methods=['GET'])
@login_required
def export_routimo():
    """
    Eksport danych do formatu Routimo EXCEL
    ZMIANA: Wszystkie rekordy OPRÓCZ wykluczonych statusów
    """
    try:
        user_email = session.get('user_email')
        reports_logger.info("Rozpoczęcie eksportu Routimo Excel", user_email=user_email)
        
        # Pobierz parametry filtrowania
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        
        # Parsuj daty
        date_from = None
        date_to = None
        
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'error': 'Nieprawidłowy format daty początkowej'}), 400
                
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'error': 'Nieprawidłowy format daty końcowej'}), 400
        
        # KLUCZOWA ZMIANA: Wykluczaj określone statusy zamiast włączać tylko transport WoodPower
        excluded_status_ids = [138625, 149779, 149778, 138624, 149777, 149763, 105114]
        
        # Buduj zapytanie SQLAlchemy
        query = BaselinkerReportOrder.query
        
        # Filtruj po dacie
        if date_from:
            query = query.filter(BaselinkerReportOrder.date_created >= date_from)
        if date_to:
            query = query.filter(BaselinkerReportOrder.date_created <= date_to)
            
        # ZMIANA: Wykluczaj statusy zamiast je włączać
        query = query.filter(~BaselinkerReportOrder.baselinker_status_id.in_(excluded_status_ids))
        
        # Sortuj po dacie i ID zamówienia
        query = query.order_by(
            BaselinkerReportOrder.date_created.desc(),
            BaselinkerReportOrder.baselinker_order_id
        )
        
        # Wykonaj zapytanie
        orders = query.all()

        reports_logger.info("Pobrano dane do eksportu Routimo Excel",
                          user_email=user_email,
                          raw_records=len(orders),
                          excluded_status_ids=excluded_status_ids,
                          date_from=date_from.isoformat() if date_from else None,
                          date_to=date_to.isoformat() if date_to else None)

        if not orders:
            return jsonify({
                'success': False,
                'error': 'Brak danych do eksportu'
            }), 400
            
        # Grupuj dane po zamówieniach
        grouped_orders = group_orders_for_routimo(orders)
        
        # ZMIANA: Generuj EXCEL zamiast CSV
        excel_content = generate_routimo_excel(grouped_orders)
        
        # ZMIANA: Przygotuj response dla Excel
        filename = f"routimo_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        response = make_response(excel_content)
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        reports_logger.info("Wygenerowano eksport Routimo Excel",
                          user_email=user_email,
                          grouped_orders=len(grouped_orders),
                          filename=filename)
        
        return response
        
    except Exception as e:
        reports_logger.error("Błąd eksportu Routimo Excel",
                           user_email=user_email if 'user_email' in locals() else 'unknown',
                           error=str(e))
        return jsonify({
            'success': False,
            'error': f'Błąd eksportu Routimo Excel: {str(e)}'
        }), 500


def generate_routimo_excel(grouped_orders):
    """
    ZAKTUALIZOWANA FUNKCJA: Generuje Excel w formacie identycznym z wzorcem
    DODANE: kolumny "Numer wew." i "Koszty kuriera netto"
    """
    # Nagłówki - ZAKTUALIZOWANE z nowymi kolumnami
    headers = [
        'Nazwa', 'Klient', 'Nazwa przesyłki', 'Numer wew.', 'Koszty kuriera netto', 'Ulica', 'Numer domu', 'Numer mieszkania',
        'Kod pocztowy', 'Miasto', 'Kraj', 'Region', 'Numer telefonu', 'Email',
        'Email klienta', 'Nip klienta', 'Początek okna czasowego', 'Koniec okna czasowego',
        'Okno czasowe', 'Czas na wykonanie zadania', 'Oczekiwana data realizacji',
        'Harmonogram', 'Pojazd', 'Typy pojazdów', 'Liczba przesyłek', 'Wielkość przesyłki',
        'Waga przesyłki', 'Wartość przesyłki', 'Forma płatności', 'Waluta',
        'Szerokość geograficzna', 'Długość geograficzna', 'Komentarz', 'Komentarz 2',
        'Uwagi', 'Dodatkowe 1', 'Dodatkowe 2'
    ]
    
    # Utwórz nowy workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    
    # Dodaj drugi pusty arkusz (jak w wzorcu)
    workbook.create_sheet("Sheet2")
    
    # STYLOWANIE NAGŁÓWKÓW - szare tło + pogrubienie + podkreślenie
    header_fill = PatternFill(
    start_color="F3F3F3",
    end_color="EFEFEF", 
    fill_type="solid"
    )

    header_font = Font(
        bold=True,
        underline='single'
    )

    header_alignment = Alignment(
        horizontal='left',        # ZMIANA: wyrównanie do lewej
        vertical='center',
        wrap_text=True           # ZMIANA: zawijanie tekstu
    )

    # OBRAMOWANIE - czarne, standardowe
    header_border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    # Dodaj nagłówki z pełnym stylowaniem
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border     # ZMIANA: dodanie obramowania
    
    # SZEROKOŚCI KOLUMN - ZAKTUALIZOWANE z nowymi kolumnami
    column_widths = {
        'A': 40.0,   # Nazwa (bardzo szeroka dla długich nazw firm)
        'B': 31.81,  # Klient  
        'C': 17.0,   # Nazwa przesyłki (ID zamówienia)
        'D': 13.0,   # NOWA: Numer wew.
        'E': 13.0,   # NOWA: Koszty kuriera netto
        'F': 32.0,   # Ulica (szeroka dla długich nazw ulic) - przesunięte z D
        'G': 9.0,    # Numer domu - przesunięte z E
        'H': 9.0,    # Numer mieszkania - przesunięte z F
        'I': 14.0,   # Kod pocztowy - przesunięte z G
        'J': 25.0,   # Miasto - przesunięte z H
        'K': 12.0,   # Kraj - przesunięte z I
        'L': 20.0,   # Region/Województwo - przesunięte z J
        'M': 15.0,   # Telefon - przesunięte z K
        'N': 25.0,   # Email (puste) - przesunięte z L
        'O': 38.0,   # Email klienta - przesunięte z M
        'P': 15.0,   # NIP (puste) - przesunięte z N
        'Q': 20.0,   # Początek okna - przesunięte z O
        'R': 20.0,   # Koniec okna - przesunięte z P
        'S': 15.0,   # Okno czasowe - przesunięte z Q
        'T': 25.0,   # Czas na zadanie - przesunięte z R
        'U': 20.0,   # Data realizacji - przesunięte z S
        'V': 15.0,   # Harmonogram - przesunięte z T
        'W': 15.0,   # Pojazd - przesunięte z U
        'X': 20.0,   # Typy pojazdów - przesunięte z V
        'Y': 15.0,   # Liczba przesyłek - przesunięte z W
        'Z': 18.0,   # Wielkość (m³) - przesunięte z X
        'AA': 15.0,  # Waga (kg) - przesunięte z Y
        'AB': 18.0,  # Wartość PLN - przesunięte z Z
        'AC': 20.0,  # Forma płatności - przesunięte z AA
        'AD': 10.0,  # Waluta - przesunięte z AB
        'AE': 20.0,  # Szerokość geo - przesunięte z AC
        'AF': 20.0,  # Długość geo - przesunięte z AD
        'AG': 70.0,  # Komentarz - przesunięte z AE
        'AH': 20.0,  # Komentarz 2 - przesunięte z AF
        'AI': 25.0,  # Uwagi - przesunięte z AG
        'AJ': 15.0,  # Dodatkowe 1 - przesunięte z AH
        'AK': 15.0   # Dodatkowe 2 - przesunięte z AI
    }
    
    # Ustaw szerokości kolumn
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # WYSOKOŚĆ WIERSZA NAGŁÓWKOWEGO - 57px jak żądasz (≈43pt)
    worksheet.row_dimensions[1].height = 43.0
    
            # Dodaj dane
    for row_idx, order in enumerate(grouped_orders, 2):  # Zaczynaj od wiersza 2
        # Wyciągnij numer domu i mieszkania z adresu
        house_number, apartment_number, clean_street = extract_house_and_apartment_number(order['delivery_address'])
        
        # Oblicz wagę (jak w oryginalnym CSV)
        weight = round(order['total_volume'] * 800, 2)
        
        # NOWE: Generuj komentarz z listą produktów (każda pozycja od nowej linii)
        products_comment = generate_products_comment_multiline(order['records'])
        
        # NOWE: Oblicz łączną liczbę sztuk wszystkich produktów w zamówieniu
        total_quantity = sum(int(record.quantity or 0) for record in order['records'])
        
        # NOWE: Oblicz koszty kuriera netto (z VAT 23%)
        delivery_cost_gross = order.get('delivery_cost', 0) or 0
        delivery_cost_net = round(float(delivery_cost_gross) / 1.23, 2) if delivery_cost_gross > 0 else 0
        
        # NOWE: Utwórz komentarz 2 z numerem Baselinker i numerem wewnętrznym
        baselinker_id = order['baselinker_order_id'] or ''
        internal_number = order.get('internal_order_number', '') or ''
        comment_2 = f"{baselinker_id}, {internal_number}" if baselinker_id and internal_number else (baselinker_id or internal_number or '')
        
        # Dane wiersza - ZAKTUALIZOWANE z nowymi kolumnami
        row_data = [
            order['customer_name'],                    # A - Nazwa
            order['customer_name'],                    # B - Klient
            order['baselinker_order_id'],              # C - Nazwa przesyłki
            order.get('internal_order_number', ''),    # D - NOWA: Numer wew.
            delivery_cost_net,                         # E - NOWA: Koszty kuriera netto
            clean_street,                              # F - Ulica (OCZYSZCZONA!) - przesunięte z D
            house_number,                              # G - Numer domu - przesunięte z E
            apartment_number,                          # H - Numer mieszkania - przesunięte z F
            order['delivery_postcode'],                # I - Kod pocztowy - przesunięte z G
            order['delivery_city'],                    # J - Miasto - przesunięte z H
            'Polska',                                  # K - Kraj - przesunięte z I
            order['delivery_state'],                   # L - Region/Województwo - przesunięte z J
            order['phone'],                            # M - Telefon - przesunięte z K
            '',                                        # N - Email (puste) - przesunięte z L
            order.get('email', ''),                    # O - Email klienta - przesunięte z M
            '',                                        # P - NIP (puste) - przesunięte z N
            '',                                        # Q - Początek okna (puste) - przesunięte z O
            '',                                        # R - Koniec okna (puste) - przesunięte z P
            '',                                        # S - Okno czasowe (puste) - przesunięte z Q
            '',                                        # T - Czas na zadanie (puste) - przesunięte z R
            '',                                        # U - Data realizacji (puste) - przesunięte z S
            '',                                        # V - Harmonogram (puste) - przesunięte z T
            '',                                        # W - Pojazd (puste) - przesunięte z U
            '',                                        # X - Typy pojazdów (puste) - przesunięte z V
            total_quantity,                            # Y - Liczba przesyłek (suma sztuk wszystkich produktów) - przesunięte z W
            round(order['total_volume'], 3),           # Z - Wielkość w m³ - przesunięte z X
            weight,                                    # AA - Waga w kg (objętość * 800) - przesunięte z Y
            round(order['order_amount_net'], 2),       # AB - Wartość w PLN - przesunięte z Z
            order.get('payment_method', ''),           # AC - Forma płatności - przesunięte z AA
            'PLN',                                     # AD - Waluta - przesunięte z AB
            '',                                        # AE - Szerokość geograficzna (puste) - przesunięte z AC
            '',                                        # AF - Długość geograficzna (puste) - przesunięte z AD
            products_comment,                          # AG - Komentarz z listą produktów (wieloliniowy) - przesunięte z AE
            comment_2,                                 # AH - Komentarz 2 (Baselinker ID, Numer wew.) - przesunięte z AF
            '',                                        # AI - Uwagi (puste) - przesunięte z AG
            '',                                        # AJ - Dodatkowe 1 (puste) - przesunięte z AH
            '',                                        # AK - Dodatkowe 2 (puste) - przesunięte z AI
        ]
        
        # Wstaw dane do wiersza
        for col_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # NOWE: Specjalne formatowanie dla kolumny komentarz (AG)
            if col_idx == 33:  # Kolumna AG - Komentarz
                cell.alignment = Alignment(
                    horizontal='left',
                    vertical='top',
                    wrap_text=True  # Zawijanie tekstu dla wieloliniowego komentarza
                )
        
        # NOWE: Automatyczne dostosowanie wysokości wiersza dla komentarza
        if products_comment and '\n' in products_comment:
            # Oszacuj liczbę linii i ustaw wysokość wiersza
            line_count = products_comment.count('\n') + 1
            row_height = max(15 * line_count, 15)  # Minimum 15pt na linię
            worksheet.row_dimensions[row_idx].height = row_height
    
    # Zapisz do BytesIO
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    reports_logger.info("Wygenerowano Excel dla Routimo z identycznym formatowaniem",
                      orders_count=len(grouped_orders))
    
    return excel_buffer.getvalue()

    
def generate_products_comment_multiline(order_records):
    """
    Generuje komentarz z listą wszystkich produktów w zamówieniu (wieloliniowy)
    Format: każda pozycja produktu od nowej linii
    "Klejonka dębowa lita A/B 200.0×30.0×3.2cm (Surowe) x1
    Klejonka bukowa lita A/B 150.0×25.0×2.8cm (Olejowana) x6"
    
    Args:
        order_records: Lista rekordów BaselinkerReportOrder dla jednego zamówienia
        
    Returns:
        str: Sformatowany wieloliniowy komentarz z produktami
    """
    if not order_records:
        return ''
    
    products_list = []
    
    for record in order_records:
        # Użyj raw_product_name z bazy danych
        product_name = record.raw_product_name or 'Produkt bez nazwy'
        quantity = int(record.quantity or 1)
        
        # Format: "Nazwa produktu x{ilość}"
        product_entry = f"{product_name} x{quantity}"
        products_list.append(product_entry)
    
    # Połącz wszystkie produkty znakiem nowej linii
    return '\n'.join(products_list)

def group_orders_for_routimo(orders):
    """
    Grupuje dane po zamówieniach dla eksportu Routimo
    Jedno zamówienie = jeden wiersz w CSV
    NOWE: Wykluczenie usług z eksportu
    DODANE: internal_order_number i delivery_cost
    
    Args:
        orders (List[BaselinkerReportOrder]): Lista rekordów z bazy danych
        
    Returns:
        List[Dict]: Lista zamówień zgrupowanych (tylko produkty fizyczne)
    """
    # NOWE: Filtruj usługi na początku - Routimo dostaje tylko produkty fizyczne
    physical_products_only = [order for order in orders if order.group_type != 'usługa']
    
    if len(orders) != len(physical_products_only):
        services_excluded = len(orders) - len(physical_products_only)
        reports_logger.info("Wykluczono usługi z eksportu Routimo",
                          total_records=len(orders),
                          physical_products=len(physical_products_only),
                          services_excluded=services_excluded)
    
    grouped = defaultdict(lambda: {
        'records': [],
        'baselinker_order_id': None,
        'internal_order_number': None,  # NOWE: dodane pole
        'customer_name': None,
        'delivery_address': None,
        'delivery_postcode': None,
        'delivery_city': None,
        'delivery_state': None,
        'phone': None,
        'email': None,
        'delivery_cost': 0,  # NOWE: dodane pole
        'payment_method': None,  # NOWE: dodane pole dla order_amount_net
        'order_amount_net': 0,  # NOWE: dodane pole
        'total_quantity': 0,
        'total_volume': 0,
        'total_value_net': 0,
        'current_status': None
    })
    
    for order in physical_products_only:
        # Klucz grupowania - baselinker_order_id lub manual_id
        if order.baselinker_order_id:
            order_key = f"bl_{order.baselinker_order_id}"
        else:
            order_key = f"manual_{order.id}"
            
        order_group = grouped[order_key]
        order_group['records'].append(order)
        
        # Ustaw dane zamówienia (z pierwszego rekordu)
        if not order_group['customer_name']:
            order_group['baselinker_order_id'] = order.baselinker_order_id or f"Manual_{order.id}"
            order_group['internal_order_number'] = order.internal_order_number or ''  # NOWE
            order_group['customer_name'] = order.customer_name or ''
            order_group['delivery_address'] = order.delivery_address or ''
            order_group['delivery_postcode'] = order.delivery_postcode or ''
            order_group['delivery_city'] = order.delivery_city or ''
            order_group['delivery_state'] = order.delivery_state or ''
            order_group['phone'] = order.phone or ''
            order_group['email'] = order.email or ''
            order_group['delivery_cost'] = float(order.delivery_cost or 0)  # NOWE
            order_group['payment_method'] = order.payment_method or ''  # NOWE
            order_group['order_amount_net'] = float(order.order_amount_net or 0)  # NOWE
            order_group['current_status'] = order.current_status or ''
        
        # Sumuj wartości - POPRAWKA: używaj właściwości SQLAlchemy
        order_group['total_quantity'] += float(order.quantity or 0)
        order_group['total_volume'] += float(order.total_volume or 0)
        order_group['total_value_net'] += float(order.value_net or 0)
    
    # Konwertuj na listę
    result = []
    for order_key, order_data in grouped.items():
        result.append(order_data)
    
    reports_logger.info("Zgrupowano zamówienia dla Routimo",
                      raw_records=len(orders),
                      grouped_orders=len(result))
    
    return result

def extract_house_and_apartment_number(address):
    """
    Wyciąga numer domu i mieszkania z adresu oraz zwraca oczyszczoną ulicę
    Obsługuje formaty: "ul. Nazwa 123", "123 Nazwa ulicy", "Nazwa 123/45"
    
    Args:
        address (str): Pełny adres
        
    Returns:
        tuple: (house_number, apartment_number, clean_street)
    """
    if not address or not isinstance(address, str):
        return '', '', address or ''
        
    original_address = address.strip()
    
    # WZORCE - NUMER PO NAZWIE ULICY (tradycyjne)
    traditional_patterns = [
        # "ul. Nazwa 123/45" 
        {
            'pattern': r'^(.+?)\s+(\d+[A-Za-z]*)\/(\d+[A-Za-z]*)$',
            'has_apartment': True,
            'street_group': 1,
            'house_group': 2,
            'apartment_group': 3
        },
        # "ul. Nazwa 123 / 45" (ze spacjami)  
        {
            'pattern': r'^(.+?)\s+(\d+[A-Za-z]*)\s*\/\s*(\d+[A-Za-z]*)$',
            'has_apartment': True,
            'street_group': 1,
            'house_group': 2,
            'apartment_group': 3
        },
        # "ul. Nazwa 123m45"
        {
            'pattern': r'^(.+?)\s+(\d+[A-Za-z]*)\s*m\.?\s*(\d+[A-Za-z]*)$',
            'has_apartment': True,
            'street_group': 1,
            'house_group': 2,
            'apartment_group': 3
        },
        # "ul. Nazwa 123" (tylko dom)
        {
            'pattern': r'^(.+?)\s+(\d+[A-Za-z]*)\s*$',
            'has_apartment': False,
            'street_group': 1,
            'house_group': 2,
            'apartment_group': None
        }
    ]
    
    # WZORCE - NUMER PRZED NAZWĄ ULICY (odwrócone)
    reversed_patterns = [
        # "123/45 Nazwa ulicy"
        {
            'pattern': r'^(\d+[A-Za-z]*)\/(\d+[A-Za-z]*)\s+(.+)$',
            'has_apartment': True,
            'street_group': 3,
            'house_group': 1,
            'apartment_group': 2
        },
        # "123 / 45 Nazwa ulicy" (ze spacjami)
        {
            'pattern': r'^(\d+[A-Za-z]*)\s*\/\s*(\d+[A-Za-z]*)\s+(.+)$',
            'has_apartment': True,
            'street_group': 3,
            'house_group': 1,
            'apartment_group': 2
        },
        # "123m45 Nazwa ulicy"
        {
            'pattern': r'^(\d+[A-Za-z]*)\s*m\.?\s*(\d+[A-Za-z]*)\s+(.+)$',
            'has_apartment': True,
            'street_group': 3,
            'house_group': 1,
            'apartment_group': 2
        },
        # "123 Nazwa ulicy" (tylko dom)
        {
            'pattern': r'^(\d+[A-Za-z]*)\s+(.+)$',
            'has_apartment': False,
            'street_group': 2,
            'house_group': 1,
            'apartment_group': None
        }
    ]
    
    # Sprawdź wszystkie wzorce - najpierw tradycyjne, potem odwrócone
    all_patterns = traditional_patterns + reversed_patterns
    
    for pattern_info in all_patterns:
        match = re.search(pattern_info['pattern'], original_address, re.IGNORECASE)
        if match:
            groups = match.groups()
            
            # Wyciągnij komponenty według grup
            street = groups[pattern_info['street_group'] - 1].strip()
            house = groups[pattern_info['house_group'] - 1].strip()
            apartment = ''
            
            if pattern_info['has_apartment'] and pattern_info['apartment_group']:
                apartment = groups[pattern_info['apartment_group'] - 1].strip()
            
            # Sprawdź czy ulica nie jest pusta
            if not street:
                continue
                
            # Oczyść ulicę delikatnie
            clean_street = clean_street_name(street)
            
            # Jeśli po czyszczeniu ulica jest pusta, spróbuj następny wzorzec
            if not clean_street:
                continue
                
            return house, apartment, clean_street
    
    # Fallback - nie znaleziono wzorca, zwróć oryginalny adres
    return '', '', original_address


def clean_street_name(street):
    """
    Delikatnie czyści nazwę ulicy z niepotrzebnych elementów
    POPRAWKA: Nie usuwa "Aleja" jeśli to część nazwy ulicy
    
    Args:
        street (str): Surowa nazwa ulicy
        
    Returns:
        str: Oczyszczona nazwa ulicy
    """
    if not street:
        return ''
    
    # Usuń zbędne białe znaki
    street = street.strip()
    
    # Usuń końcowe przecinki i kropki
    street = re.sub(r'[,\.]+$', '', street).strip()
    
    # Usuń miasto z początku (tylko jeśli po przecinku jest coś więcej)
    # "Warszawa, ul. Nowa" → "ul. Nowa"
    city_pattern = r'^([A-ZĄĆĘŁŃÓŚŹŻ][a-ząćęłńóśźż]+)\s*,\s*(.+)$'
    city_match = re.match(city_pattern, street)
    if city_match and city_match.group(2).strip():
        street = city_match.group(2).strip()
    
    # POPRAWKA: Usuń prefiksy TYLKO jeśli są na początku i po nich jest jeszcze tekst
    # ALE zachowaj "Aleja Nazwa" jako całość - nie traktuj "Aleja" jako prefiksu do usunięcia
    
    # Lista prefixów do usunięcia TYLKO jeśli są na samym początku
    prefixes_to_remove = ['ul', 'ulica']  # Skróciłem listę!
    
    for prefix in prefixes_to_remove:
        # Usuń tylko "ul." lub "ulica" na początku, ale zostaw "al.", "pl.", "os."
        pattern = rf'^{prefix}\.?\s+(.+)$'
        match = re.match(pattern, street, re.IGNORECASE)
        if match and match.group(1).strip():
            street = match.group(1).strip()
            break  # Usuń tylko pierwszy pasujący prefiks
    
    return street

def generate_routimo_csv(grouped_orders):
    """
    Generuje CSV w formacie Routimo
    """
    # Nagłówki pozostają bez zmian...
    headers = [
        'Nazwa', 'Klient', 'Nazwa przesyłki', 'Ulica', 'Numer domu', 'Numer mieszkania',
        'Kod pocztowy', 'Miasto', 'Kraj', 'Region', 'Numer telefonu', 'Email',
        'Email klienta', 'Nip klienta', 'Początek okna czasowego', 'Koniec okna czasowego',
        'Okno czasowe', 'Czas na wykonanie zadania', 'Oczekiwana data realizacji',
        'Harmonogram', 'Pojazd', 'Typy pojazdów', 'Liczba przesyłek', 'Wielkość przesyłki',
        'Waga przesyłki', 'Wartość przesyłki', 'Forma płatności', 'Waluta',
        'Szerokość geograficzna', 'Długość geograficzna', 'Komentarz', 'Komentarz 2',
        'Uwagi', 'Dodatkowe 1', 'Dodatkowe 2'
    ]
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    
    for order in grouped_orders:
        # ZMIANA: Używaj nowej funkcji z czyszczeniem ulicy
        house_number, apartment_number, clean_street = extract_house_and_apartment_number(order['delivery_address'])
        
        # Oblicz wagę
        weight = round(order['total_volume'] * 800, 2)
        
        row = [
            order['customer_name'],                    # A - Nazwa
            order['customer_name'],                    # B - Klient
            order['baselinker_order_id'],              # C - Nazwa przesyłki
            clean_street,                              # D - Ulica (OCZYSZCZONA!)
            house_number,                              # E - Numer domu
            apartment_number,                          # F - Numer mieszkania
            order['delivery_postcode'],                # G - Kod pocztowy
            order['delivery_city'],                    # H - Miasto
            'Polska',                                  # I - Kraj
            order['delivery_state'],                   # J - Region
            order['phone'],                            # K - Numer telefonu
            '',                                        # L - Email (puste)
            order['email'],                            # M - Email klienta
            '',                                        # N - Nip klienta (puste)
            '',                                        # O - Początek okna czasowego (puste)
            '',                                        # P - Koniec okna czasowego (puste)
            '',                                        # Q - Okno czasowe (puste)
            '',                                        # R - Czas na wykonanie zadania (puste)
            '',                                        # S - Oczekiwana data realizacji (puste)
            '',                                        # T - Harmonogram (puste)
            '',                                        # U - Pojazd (puste)
            '',                                        # V - Typy pojazdów (puste)
            int(order['total_quantity']),              # W - Liczba przesyłek
            round(order['total_volume'], 4),           # X - Wielkość przesyłki (m³)
            weight,                                    # Y - Waga przesyłki (kg)
            round(order['total_value_net'], 2),        # Z - Wartość przesyłki
            '',                                        # AA - Forma płatności (puste)
            'PLN',                                     # AB - Waluta
            '',                                        # AC - Szerokość geograficzna (puste)
            '',                                        # AD - Długość geograficzna (puste)
            '',                                        # AE - Komentarz (puste)
            '',                                        # AF - Komentarz 2 (puste)
            '',                                        # AG - Uwagi (puste)
            '',                                        # AH - Dodatkowe 1 (puste)
            '',                                        # AI - Dodatkowe 2 (puste)
        ]
        
        writer.writerow(row)
    
    csv_content = output.getvalue()
    output.close()
    
    reports_logger.info("Wygenerowano CSV dla Routimo z oczyszczonymi adresami",
                      orders_count=len(grouped_orders))
    
    return csv_content

def extract_volume_from_product_name(product_name: str) -> Optional[float]:
    """
    Wyodrębnia objętość z nazwy produktu w różnych formatach.
    """
    if not product_name:
        return None
    
    # Wzorce dla różnych formatów objętości
    volume_patterns = [
        r'(\d+[,.]?\d*)\s*[mM]3?\b',
        r'(\d+[,.]?\d*)\s*[mM]³\b',
        r'(\d+[,.]?\d*)\s*[mM]\s*3\b',
        r'\(\s*(\d+[,.]?\d*)\s*[mM]3?\s*\)',
    ]
    
    for pattern in volume_patterns:
        matches = re.findall(pattern, product_name, re.IGNORECASE)
        if matches:
            volume_str = matches[0].replace(',', '.')
            try:
                volume = float(volume_str)
                print(f"[DEBUG] Found volume {volume} m³ in '{product_name}'")
                return volume
            except ValueError:
                continue
    
    return None

def extract_wood_species_from_product_name(product_name: str) -> Optional[str]:
    """
    Wyodrębnia gatunek drewna z nazwy produktu.
    
    Args:
        product_name (str): Nazwa produktu
        
    Returns:
        Optional[str]: Gatunek drewna lub None
    """
    if not product_name:
        return None
    
    name_lower = product_name.lower()
    
    # Mapowanie różnych form nazw gatunków
    species_mapping = {
        'dąb': ['dąb', 'dab', 'dębowy', 'dębowa', 'dębowe'],
        'buk': ['buk', 'bukowy', 'bukowa', 'bukowe', 'bukowych'],
        'jesion': ['jesion', 'jesionowy', 'jesionowa', 'jesionowe'],
        'sosna': ['sosna', 'sosnowy', 'sosnowa', 'sosnowe'],
        'brzoza': ['brzoza', 'brzozowy', 'brzozowa', 'brzozowe'],
        'wiąz': ['wiąz', 'wiązowy', 'wiązowa', 'wiązowe'],
        'klon': ['klon', 'klonowy', 'klonowa', 'klonowe'],
    }
    
    for standard_name, variants in species_mapping.items():
        for variant in variants:
            if variant in name_lower:
                return standard_name
    
    print(f"[DEBUG] Nie znaleziono gatunku w '{product_name}'")
    return None

def extract_technology_from_product_name(product_name: str) -> Optional[str]:
    """
    Wyodrębnia technologię z nazwy produktu.
    
    Args:
        product_name (str): Nazwa produktu
        
    Returns:
        Optional[str]: Technologia lub None
    """
    if not product_name:
        return None
    
    name_lower = product_name.lower()
    
    # Mapowanie różnych form technologii
    technology_mapping = {
        'lity': ['lity', 'lite', 'litych', 'litej', 'litego'],
        'mikrowczep': ['mikrowczep', 'micro', 'wczep'],
        'klejony': ['klejony', 'klejona', 'klejone', 'klejonych'],
        'fornir': ['fornir', 'fornirowany', 'fornirowana'],
    }
    
    for standard_name, variants in technology_mapping.items():
        for variant in variants:
            if variant in name_lower:
                return standard_name
    
    return None

def extract_wood_class_from_product_name(product_name: str) -> Optional[str]:
    """
    Wyodrębnia klasę drewna z nazwy produktu.
    
    Args:
        product_name (str): Nazwa produktu
        
    Returns:
        Optional[str]: Klasa drewna lub None
    """
    if not product_name:
        return None
    
    # Wzorce dla klas drewna
    class_patterns = [
        r'\b([AB]/[AB])\b',  # A/B, B/B
        r'\b([AB]-[AB])\b',  # A-B, B-B  
        r'\bklasa\s+([AB]/[AB])\b',  # klasa A/B
        r'\bklasa\s+([AB]-[AB])\b',  # klasa A-B
    ]
    
    for pattern in class_patterns:
        matches = re.findall(pattern, product_name, re.IGNORECASE)
        if matches:
            wood_class = matches[0].upper().replace('-', '/')  # Normalizuj do formatu A/B
            return wood_class
    
    print(f"[DEBUG] Nie znaleziono klasy w '{product_name}'")
    return None

def analyze_product_for_volume_and_attributes(product_name):
    if not product_name:
        return {
            'has_dimensions': False,
            'has_volume': False,
            'volume': None,
            'wood_species': None,
            'technology': None,
            'wood_class': None,
            'analysis_type': 'empty'
        }
    
    # Sprawdź wymiary
    has_dimensions = check_product_dimensions(product_name)
    
    # Sprawdź objętość
    volume = extract_volume_from_product_name(product_name)
    has_volume = volume is not None
    
    # Wyodrębnij atrybuty
    wood_species = extract_wood_species_from_product_name(product_name)
    technology = extract_technology_from_product_name(product_name)
    wood_class = extract_wood_class_from_product_name(product_name)
    
    # NOWA LOGIKA: Sprawdź typ produktu z nazwy
    from .parser import ProductNameParser
    parser = ProductNameParser()
    parsed = parser.parse_product_name(product_name)
    product_type = parsed.get('product_type')
    
    # OKREŚL ANALYSIS_TYPE na podstawie typu produktu
    analysis_type = 'unknown'
    
    # WORKI OPAŁOWE: Nie wymagają walidacji objętości ani wymiarów
    if product_type == 'worek opałowy':
        analysis_type = 'no_validation_needed'  # Nowy typ - bez walidacji
    
    # SUSZENIE: Wymaga walidacji objętości (podobnie jak produkty bez wymiarów)
    elif product_type == 'suszenie':
        if has_volume:
            analysis_type = 'volume_only'
        else:
            analysis_type = 'manual_input_needed'
    
    # TARCICA: Wymaga walidacji objętości (podobnie jak produkty bez wymiarów)
    elif product_type == 'tarcica':
        if has_volume:
            analysis_type = 'volume_only'
        else:
            analysis_type = 'manual_input_needed'
    
    # ISTNIEJĄCE PRODUKTY (klejonka, deska): dotychczasowa logika
    else:
        # Sprawdź czy wymiary to rzeczywiste wymiary (nie heurystyka)
        has_real_dimensions = False
        if has_dimensions:
            import re
            real_dimension_patterns = [
                r'\d+[,.]?\d*\s*x\s*\d+[,.]?\d*\s*x\s*\d+[,.]?\d*',  # 200,4x89x4.5 (3 wymiary)
                r'\d+[,.]?\d*\s*×\s*\d+[,.]?\d*\s*×\s*\d+[,.]?\d*',  # 200,4×89×4.5
                r'\d+[,.]?\d*\s*x\s*\d+[,.]?\d*(?!\s*x)',  # 200,4x89 (2 wymiary)
                r'\d+[,.]?\d*\s*×\s*\d+[,.]?\d*(?!\s*×)',  # 200,4×89
            ]
            
            for pattern in real_dimension_patterns:
                if re.search(pattern, product_name):
                    has_real_dimensions = True
                    break
                    
        # LOGIKA PRIORYTETU dla klejonki/deski:
        if has_real_dimensions and not has_volume:
            analysis_type = 'dimensions_priority'
        elif has_real_dimensions and has_volume:
            analysis_type = 'dimensions_priority'  # Rzeczywiste wymiary wygrywają z objętością
        elif not has_real_dimensions and has_volume:
            analysis_type = 'volume_only'  # Objętość wygrywa z heurystyką wymiarów
        elif not has_real_dimensions and not has_volume and has_dimensions:
            analysis_type = 'manual_input_needed'  # Tylko heurystyka wymiarów - lepiej zapytać użytkownika
        else:
            analysis_type = 'manual_input_needed'
    
    result = {
        'has_dimensions': has_dimensions,
        'has_volume': has_volume,
        'volume': volume,
        'wood_species': wood_species,
        'technology': technology,
        'wood_class': wood_class,
        'analysis_type': analysis_type,
        'product_type': product_type  # NOWE: dodaj product_type do rezultatu
    }
    return result

def should_show_volume_modal_for_orders(orders_data: list) -> Tuple[bool, list]:
    """
    Sprawdza czy któreś z zamówień wymaga modala objętości.
    
    Args:
        orders_data (list): Lista danych zamówień
        
    Returns:
        Tuple[bool, list]: (czy_pokazać_modal, lista_produktów_wymagających_objętości)
    """
    products_needing_volume = []
    
    for order in orders_data:
        for product in order.get('products', []):
            analysis = analyze_product_for_volume_and_attributes(product.get('name', ''))
            
            # Jeśli produkt nie ma wymiarów ani objętości, wymaga ręcznego wprowadzenia
            if analysis['analysis_type'] == 'manual_input_needed':
                products_needing_volume.append({
                    'order_id': order.get('order_id'),
                    'product_name': product.get('name'),
                    'quantity': product.get('quantity', 1),
                    'analysis': analysis
                })
    
    should_show_modal = len(products_needing_volume) > 0
    print(f"[DEBUG] Modal objętości: {should_show_modal}, produktów do uzupełnienia: {len(products_needing_volume)}")
    
    return should_show_modal, products_needing_volume

@reports_bp.route('/api/save-orders-with-volumes', methods=['POST'])
@login_required
def api_save_orders_with_volumes():
    """
    NOWY ENDPOINT: Zapisuje zamówienia z uzupełnionymi objętościami i atrybutami
    """
    user_email = session.get('user_email')
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': 'Brak danych w żądaniu'
            }), 400

        order_ids = data.get('order_ids', [])
        volume_fixes = data.get('volume_fixes', {})
        orders_data = data.get('orders_data', [])

        if not order_ids:
            return jsonify({
                'success': False,
                'error': 'Brak ID zamówień do przetworzenia'
            }), 400

        # ✅ KONWERSJA: Przekonwertuj order_ids na integery dla spójności
        try:
            order_ids_int = [int(order_id) for order_id in order_ids]
        except (ValueError, TypeError) as e:
            return jsonify({
                'success': False,
                'error': 'Błędny format ID zamówień'
            }), 400

        reports_logger.info("Rozpoczęcie zapisywania zamówień z objętościami",
                          user_email=user_email,
                          order_ids_count=len(order_ids_int),
                          orders_data_count=len(orders_data),
                          volume_fixes_count=len(volume_fixes))

        # Sprawdź czy wybrane zamówienia już istnieją w bazie
        service = get_reports_service()
        existing_order_ids = service.get_existing_order_ids(order_ids_int)  # ✅ Użyj int wersji
        
        # Filtruj tylko nowe zamówienia
        new_order_ids = [order_id for order_id in order_ids_int if order_id not in existing_order_ids]

        if not new_order_ids:
            return jsonify({
                'success': True,
                'message': 'Wszystkie wybrane zamówienia już istnieją w bazie danych',
                'orders_added': 0,
                'orders_skipped': len(order_ids)
            })

        # ✅ POPRAWKA: Filtruj orders_data porównując integery z integerami
        filtered_orders_data = [
            order for order in orders_data 
            if order.get('order_id') in new_order_ids  # Teraz oba są integerami!
        ]

        if not filtered_orders_data:
            reports_logger.warning("Brak danych zamówień po filtrowaniu",
                                 user_email=user_email,
                                 original_orders_data_count=len(orders_data),
                                 new_order_ids=new_order_ids,
                                 existing_order_ids=list(existing_order_ids))
            return jsonify({
                'success': False,
                'error': 'Brak danych zamówień do przetworzenia'
            }), 400

        # Zastosuj poprawki objętości jeśli zostały podane
        if volume_fixes:
            service.set_volume_fixes(volume_fixes)
            reports_logger.info("🔍 DEBUGGING VOLUME FIXES:")
            reports_logger.info(f"📊 Liczba kluczy w volume_fixes: {len(volume_fixes)}")
            for product_key, fixes in volume_fixes.items():
                reports_logger.info(f"🔑 Klucz: {product_key}")
                reports_logger.info(f"📦 Dane: {fixes}")
                reports_logger.info(f"🔢 Objętość: {fixes.get('volume', 'BRAK')}")
                reports_logger.info(f"🌳 Gatunek: {fixes.get('wood_species', 'BRAK')}")
                reports_logger.info(f"🔧 Technologia: {fixes.get('technology', 'BRAK')}")
                reports_logger.info(f"📏 Klasa: {fixes.get('wood_class', 'BRAK')}")
    
            # ✅ SPRAWDŹ CZY SERVICE MA DOSTĘP DO DANYCH
            reports_logger.info("🔍 SPRAWDZENIE SERVICE:")
            reports_logger.info(f"📊 service.volume_fixes keys: {list(service.volume_fixes.keys()) if hasattr(service, 'volume_fixes') else 'BRAK ATRYBUTU'}")
    
            # ✅ TESTUJ get_volume_fix_attribute
            for product_key in volume_fixes.keys():
                test_volume = service.get_volume_fix_attribute(product_key, 'volume')
                test_species = service.get_volume_fix_attribute(product_key, 'wood_species')
                reports_logger.info(f"🧪 TEST get_volume_fix_attribute dla {product_key}:")
                reports_logger.info(f"   📦 volume: {test_volume}")
                reports_logger.info(f"   🌳 wood_species: {test_species}")
        
        else:
            reports_logger.info("⚠️ BRAK volume_fixes - używamy automatycznej analizy")

        reports_logger.info("Ustawiono poprawki objętości dla {} produktów".format(len(volume_fixes) if volume_fixes else 0))

        # ✅ DODAJ DEBUG PRZED WYWOŁANIEM FUNKCJI ZAPISUJĄCEJ
        reports_logger.info("🔍 DEBUG PRZED WYWOŁANIEM FUNKCJI ZAPISUJĄCEJ:")
        reports_logger.info(f"   📊 new_order_ids: {new_order_ids}")
        reports_logger.info(f"   📦 filtered_orders_data count: {len(filtered_orders_data)}")
        reports_logger.info(f"   🔧 service.volume_fixes keys: {list(service.volume_fixes.keys()) if hasattr(service, 'volume_fixes') else 'BRAK'}")
        
        # Sprawdź pierwszy produkt w pierwszym zamówieniu
        if filtered_orders_data and len(filtered_orders_data) > 0:
            first_order = filtered_orders_data[0]
            reports_logger.info(f"   🎯 Pierwszy order_id: {first_order.get('order_id')}")
            first_products = first_order.get('products', [])
            if first_products:
                first_product = first_products[0]
                reports_logger.info(f"   📝 Pierwszy produkt: {first_product.get('name', 'BRAK NAZWY')}")
                product_id = first_product.get('product_id', 'unknown')
                expected_key = f"{first_order.get('order_id')}_{product_id}"
                reports_logger.info(f"   🔑 Oczekiwany klucz: {expected_key}")
                
                # ✅ DODAJ DEBUG STRUKTURY PRODUKTU
                reports_logger.info(f"   🔍 Struktura pierwszego produktu:")
                reports_logger.info(f"   📦 product_id: '{first_product.get('product_id', 'BRAK')}'")
                reports_logger.info(f"   📝 name: '{first_product.get('name', 'BRAK')}'")
                reports_logger.info(f"   🔢 quantity: {first_product.get('quantity', 'BRAK')}")
                reports_logger.info(f"   💰 price_brutto: {first_product.get('price_brutto', 'BRAK')}")
                
                # Sprawdź czy klucz istnieje w volume_fixes
                if hasattr(service, 'volume_fixes') and expected_key in service.volume_fixes:
                    fix_data = service.volume_fixes[expected_key]
                    reports_logger.info(f"   ✅ Klucz znaleziony w volume_fixes: {fix_data}")
                else:
                    reports_logger.info(f"   ❌ Klucz NIE ZNALEZIONY w volume_fixes!")
                    
                    # Sprawdź wszystkie dostępne klucze
                    if hasattr(service, 'volume_fixes'):
                        available_keys = list(service.volume_fixes.keys())
                        reports_logger.info(f"   🔍 Dostępne klucze: {available_keys}")
                        
                    # ✅ SPRAWDŹ WSZYSTKIE PRODUKTY W ZAMÓWIENIU
                    reports_logger.info(f"   📋 Wszystkie produkty w zamówieniu:")
                    for idx, prod in enumerate(first_products):
                        order_product_id = prod.get('order_product_id')
                        product_id_raw = prod.get('product_id')
                        product_index = prod.get('product_index', idx)  # ✅ UŻYJ product_index Z DANYCH LUB idx
                        prod_name = prod.get('name', 'BRAK NAZWY')
            
                        # ✅ UŻYJ POPRAWIONEJ FUNKCJI generate_product_key_router
                        key_for_this_prod = generate_product_key_router(first_order.get('order_id'), prod, product_index)
            
                        has_volume_data = key_for_this_prod in service.volume_fixes if hasattr(service, 'volume_fixes') else False
            
                        reports_logger.info(f"      {idx+1}. order_product_id: '{order_product_id}' | product_id: '{product_id_raw}' | product_index: {product_index} | Nazwa: '{prod_name}' | Klucz: '{key_for_this_prod}' | Ma dane: {has_volume_data}")

        # Przekaż przefiltrowane dane zamówień
        result = _sync_selected_orders_with_volume_analysis(service, new_order_ids, filtered_orders_data)
        
        # ✅ DODAJ DEBUG PO WYWOŁANIU FUNKCJI ZAPISUJĄCEJ
        reports_logger.info("🔍 DEBUG PO WYWOŁANIU FUNKCJI ZAPISUJĄCEJ:")
        reports_logger.info(f"   📊 Result success: {result.get('success')}")
        reports_logger.info(f"   📈 Orders processed: {result.get('orders_processed')}")
        reports_logger.info(f"   📝 Orders added: {result.get('orders_added')}")
        if not result.get('success'):
            reports_logger.info(f"   ❌ Error: {result.get('error')}")

        # ✅ POPRAWKA: Wyczyść poprawki objętości DOPIERO PO zakończeniu zapisu
        # (nie wcześniej, bo _sync_selected_orders_with_volume_analysis może jeszcze ich używać!)

        if result.get('success'):
            # ✅ TUTAJ jest właściwe miejsce na czyszczenie volume_fixes
            if volume_fixes:
                service.clear_volume_fixes()
                
            reports_logger.info("Zapisywanie zamówień z objętościami zakończone pomyślnie",
                              orders_processed=result.get('orders_processed', 0),
                              orders_added=result.get('orders_added', 0))
            return jsonify(result)
        else:
            # ✅ W przypadku błędu też wyczyść volume_fixes
            if volume_fixes:
                service.clear_volume_fixes()
                
            return jsonify(result), 500
            
    except Exception as e:
        reports_logger.error("Błąd zapisywania zamówień z objętościami",
                           user_email=user_email,
                           error=str(e))
        return jsonify({
            'success': False,
            'error': f'Błąd zapisywania zamówień: {str(e)}'
        }), 500

def _sync_selected_orders_with_volume_analysis(service, order_ids, orders_data):
    """
    FUNKCJA POMOCNICZA: Synchronizuje wybrane zamówienia z uwzględnieniem analizy objętości
    """
    try:
        reports_logger.info("Rozpoczęcie synchronizacji z analizą objętości", 
                          orders_count=len(order_ids))

        # NIE POBIERAJ Z API - użyj przesłanych danych
        if not orders_data:
            return {
                'success': False,
                'error': 'Brak danych zamówień do przetworzenia'
            }

        # Przetwórz zamówienia z analizą objętości
        orders_added = 0
        orders_processed = 0
        processing_errors = []

        for order_data in orders_data:
            try:
                reports_logger.info("Zapisywanie zamówienia z analizą objętości",
                                  order_id=order_data.get('order_id'),
                                  products_count=len(order_data.get('products', [])))
                
                # ✅ POPRAWKA: Zamiast _convert_order_to_records, użyj indywidualnego przetwarzania
                # które respektuje product_index z frontendu
                
                order_id = order_data.get('order_id')
                products = order_data.get('products', [])
                
                if not products:
                    reports_logger.warning(f"Zamówienie {order_id} nie ma produktów")
                    continue
                
                saved_records = []
                
                # ✅ PRZETWARZAJ KAŻDY PRODUKT INDYWIDUALNIE Z ZACHOWANIEM product_index
                for product_index, product in enumerate(products):
                    try:
                        # ✅ POPRAWKA: Użyj product_index z frontendu jeśli dostępny, w przeciwnym razie enumerate
                        frontend_product_index = product.get('product_index')
                        if frontend_product_index is not None:
                            actual_product_index = frontend_product_index
                            reports_logger.debug(f"Używam product_index z frontendu: {actual_product_index} dla produktu {product.get('name', 'unknown')}")
                        else:
                            actual_product_index = product_index
                            reports_logger.warning(f"Brak product_index dla produktu {product.get('name', 'unknown')} w zamówieniu {order_id} - używam enumerate: {actual_product_index}")
                        
                        # ✅ UŻYJ FUNKCJI Z service.py KTÓRA UŻYWA product_index
                        record_data = service.prepare_order_record_data_with_volume_analysis(
                            order_data, product, actual_product_index
                        )
                        
                        # Zapisz rekord
                        record = service.create_report_record(record_data)
                        saved_records.append(record)
                        
                        reports_logger.debug(f"Zapisano produkt: {product.get('name', 'unknown')} z indeksem {actual_product_index}")
                        
                    except Exception as e:
                        error_msg = f"Błąd zapisywania produktu {product.get('name', 'unknown')}: {str(e)}"
                        reports_logger.error(error_msg)
                        processing_errors.append({
                            'order_id': order_id,
                            'product_name': product.get('name', 'unknown'),
                            'error': str(e)
                        })
                        continue
                
                if saved_records:
                    # Commit dla całego zamówienia
                    db.session.commit()
                    orders_added += len(saved_records)
                    
                    # Oblicz łączną objętość dla logowania
                    total_volume = sum(float(r.total_volume or 0) for r in saved_records)
                    avg_price = saved_records[0].avg_order_price_per_m3 if saved_records else 0.0
                    
                    reports_logger.info("Zapisano rekord zamówienia z objętością", 
                                      order_id=order_id,
                                      products_count=len(saved_records),
                                      avg_order_price_per_m3=avg_price,
                                      total_volume=total_volume)
                    orders_processed += 1
                else:
                    reports_logger.warning(f"Brak zapisanych produktów dla zamówienia {order_id}")
                        
            except Exception as e:
                db.session.rollback()  # ✅ DODAJ ROLLBACK W PRZYPADKU BŁĘDU
                error_msg = f"Błąd przetwarzania zamówienia {order_data.get('order_id', 'unknown')}: {str(e)}"
                processing_errors.append({
                    'order_id': order_data.get('order_id', 'unknown'),
                    'error': str(e)
                })
                reports_logger.error(error_msg)
                continue

        # Przygotuj wynik
        result = {
            'success': True,
            'orders_processed': orders_processed,
            'orders_added': orders_added,
            'message': f'Pomyślnie zapisano {orders_processed} zamówień. Dodano: {orders_added} pozycji.'
        }
        
        if processing_errors:
            result['warnings'] = processing_errors
            result['message'] += f' Błędów: {len(processing_errors)}.'
        
        return result
        
    except Exception as e:
        db.session.rollback()  # ✅ DODAJ ROLLBACK W PRZYPADKU KRYTYCZNEGO BŁĘDU
        reports_logger.error("Krytyczny błąd synchronizacji z analizą objętości", error=str(e))
        return {
            'success': False,
            'error': f'Krytyczny błąd synchronizacji: {str(e)}'
        }


@reports_bp.route('/api/map-statistics', methods=['GET'])
@login_required
def api_map_statistics():
    """
    API endpoint dla danych mapy województw
    Zwraca statystyki produkcji pogrupowane według województw
    """
    try:
        user_email = session.get('user_email', 'Nieznany użytkownik')
        reports_logger.info("Żądanie danych mapy województw", user_email=user_email)
        
        # Pobierz filtry z URL (opcjonalnie)
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        
        # Parsuj daty jeśli podane
        date_from = None
        date_to = None
        
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except ValueError:
                reports_logger.warning("Nieprawidłowy format date_from", 
                                     date_from=date_from_str, user_email=user_email)
        
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                reports_logger.warning("Nieprawidłowy format date_to", 
                                     date_to=date_to_str, user_email=user_email)
        
        # Pobierz dane z bazy
        query = BaselinkerReportOrder.get_filtered_orders(
            date_from=date_from,
            date_to=date_to
        )
        
        orders = query.all()
        
        reports_logger.info("Pobrano zamówienia dla mapy", 
                          count=len(orders), 
                          user_email=user_email,
                          date_from=date_from.isoformat() if date_from else "wszystkie",
                          date_to=date_to.isoformat() if date_to else "wszystkie")
        
        # Statusy dla grupowania
        IN_PRODUCTION_STATUSES = [155824, 138619, 148830, 148831, 148832]  # W produkcji
        READY_STATUSES = [138620, 138623, 149777]  # Wyprodukowane
        
        # Inicjalizuj dane dla wszystkich 16 województw
        voivodeships_data = {}
        all_voivodeships = [
            'dolnośląskie', 'kujawsko-pomorskie', 'lubelskie', 'lubuskie',
            'łódzkie', 'małopolskie', 'mazowieckie', 'opolskie',
            'podkarpackie', 'podlaskie', 'pomorskie', 'śląskie',
            'świętokrzyskie', 'warmińsko-mazurskie', 'wielkopolskie',
            'zachodniopomorskie'
        ]
        
        # Inicjalizuj wszystkie województwa z zerami
        for voivodeship in all_voivodeships:
            voivodeships_data[voivodeship] = {
                'production_volume': 0.0,
                'ready_volume': 0.0,
                'total_volume': 0.0,
                'production_value_net': 0.0,
                'ready_value_net': 0.0,
                'total_value_net': 0.0,
                'orders_count': 0
            }
        
        # Mapowanie nazw województw do kluczy (normalizacja)
        voivodeship_mapping = {
            'dolnośląskie': 'dolnośląskie',
            'dolnoslaskie': 'dolnośląskie',
            'kujawsko-pomorskie': 'kujawsko-pomorskie',
            'kujawsko pomorskie': 'kujawsko-pomorskie',
            'lubelskie': 'lubelskie',
            'lubuskie': 'lubuskie', 
            'łódzkie': 'łódzkie',
            'lódzkie': 'łódzkie',
            'lodzkie': 'łódzkie',
            'małopolskie': 'małopolskie',
            'malopolskie': 'małopolskie',
            'mazowieckie': 'mazowieckie',
            'opolskie': 'opolskie',
            'podkarpackie': 'podkarpackie',
            'podlaskie': 'podlaskie',
            'pomorskie': 'pomorskie',
            'śląskie': 'śląskie',
            'slaskie': 'śląskie',
            'sląskie': 'śląskie',
            'świętokrzyskie': 'świętokrzyskie',
            'swietokrzyskie': 'świętokrzyskie',
            'świętokrzyskie': 'świętokrzyskie',
            'warmińsko-mazurskie': 'warmińsko-mazurskie',
            'warminsko-mazurskie': 'warmińsko-mazurskie',
            'warmińsko mazurskie': 'warmińsko-mazurskie',
            'wielkopolskie': 'wielkopolskie',
            'zachodniopomorskie': 'zachodniopomorskie'
        }
        
        # Grupuj dane według województw
        for order in orders:
            delivery_state = order.delivery_state
            if not delivery_state:
                continue
                
            # Normalizuj nazwę województwa
            state_normalized = delivery_state.lower().strip()
            voivodeship_key = voivodeship_mapping.get(state_normalized)
            
            if not voivodeship_key:
                reports_logger.debug("Nieznane województwo", 
                                   delivery_state=delivery_state,
                                   user_email=user_email)
                continue
            
            # Pobierz wartości
            volume = float(order.total_volume or 0)
            value_net = float(order.value_net or 0)
            status_id = order.baselinker_status_id
            
            # Zwiększ licznik zamówień
            voivodeships_data[voivodeship_key]['orders_count'] += 1
            
            # Klasyfikuj według statusu
            if status_id in IN_PRODUCTION_STATUSES:
                voivodeships_data[voivodeship_key]['production_volume'] += volume
                voivodeships_data[voivodeship_key]['production_value_net'] += value_net
            elif status_id in READY_STATUSES:
                voivodeships_data[voivodeship_key]['ready_volume'] += volume
                voivodeships_data[voivodeship_key]['ready_value_net'] += value_net
        
        # Oblicz łączne wartości
        for voivodeship_key in voivodeships_data:
            data = voivodeships_data[voivodeship_key]
            data['total_volume'] = data['production_volume'] + data['ready_volume']
            data['total_value_net'] = data['production_value_net'] + data['ready_value_net']
        
        # Mapowanie kluczy dla frontendu (z polskich do angielskich ID)
        frontend_mapping = {
            'dolnośląskie': 'dolnoslaskie',
            'kujawsko-pomorskie': 'kujawsko-pomorskie', 
            'lubelskie': 'lubelskie',
            'lubuskie': 'lubuskie',
            'łódzkie': 'lodzkie',
            'małopolskie': 'malopolskie',
            'mazowieckie': 'mazowieckie',
            'opolskie': 'opolskie',
            'podkarpackie': 'podkarpackie',
            'podlaskie': 'podlaskie',
            'pomorskie': 'pomorskie',
            'śląskie': 'slaskie',
            'świętokrzyskie': 'swietokrzyskie',
            'warmińsko-mazurskie': 'warminsko-mazurskie',
            'wielkopolskie': 'wielkopolskie',
            'zachodniopomorskie': 'zachodniopomorskie'
        }
        
        # Przekonwertuj klucze dla frontendu
        frontend_data = {}
        for polish_key, data in voivodeships_data.items():
            frontend_key = frontend_mapping.get(polish_key, polish_key)
            frontend_data[frontend_key] = data
        
        # Oblicz statystyki ogólne
        total_production_volume = sum(data['production_volume'] for data in voivodeships_data.values())
        total_ready_volume = sum(data['ready_volume'] for data in voivodeships_data.values())
        total_volume = sum(data['total_volume'] for data in voivodeships_data.values())
        total_orders = sum(data['orders_count'] for data in voivodeships_data.values())
        
        result = {
            'status': 'success',
            'data': frontend_data,
            'summary': {
                'total_production_volume': total_production_volume,
                'total_ready_volume': total_ready_volume,
                'total_volume': total_volume,
                'total_orders': total_orders,
                'processed_orders': len(orders),
                'date_from': date_from.isoformat() if date_from else None,
                'date_to': date_to.isoformat() if date_to else None
            }
        }
        
        reports_logger.info("Zwrócono dane mapy", 
                          summary=result['summary'],
                          user_email=user_email)
        
        return jsonify(result)
        
    except Exception as e:
        reports_logger.error("Błąd API map-statistics", 
                           error=str(e), 
                           user_email=session.get('user_email', 'Nieznany'))
        return jsonify({
            'status': 'error',
            'message': 'Błąd podczas pobierania danych mapy',
            'error': str(e)
        }), 500
    """
    API endpoint dla danych mapy województw
    Zwraca statystyki produkcji pogrupowane według województw
    """
    try:
        user_email = session.get('user_email', 'Nieznany użytkownik')
        reports_logger.info("Żądanie danych mapy województw", user_email=user_email)
        
        # Pobierz filtry z URL (opcjonalnie)
        date_from_str = request.args.get('date_from')
        date_to_str = request.args.get('date_to')
        
        # Parsuj daty jeśli podane
        date_from = None
        date_to = None
        
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except ValueError:
                reports_logger.warning("Nieprawidłowy format date_from", 
                                     date_from=date_from_str, user_email=user_email)
        
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                reports_logger.warning("Nieprawidłowy format date_to", 
                                     date_to=date_to_str, user_email=user_email)
        
        # Pobierz dane z bazy
        query = BaselinkerReportOrder.get_filtered_orders(
            date_from=date_from,
            date_to=date_to
        )
        
        orders = query.all()
        
        reports_logger.info("Pobrano zamówienia dla mapy", 
                          count=len(orders), 
                          user_email=user_email,
                          date_from=date_from.isoformat() if date_from else "wszystkie",
                          date_to=date_to.isoformat() if date_to else "wszystkie")
        
        # Statusy dla grupowania
        IN_PRODUCTION_STATUSES = [155824, 138619, 148830, 148831, 148832]  # W produkcji
        READY_STATUSES = [138620, 138623, 149777]  # Wyprodukowane
        
        # Inicjalizuj dane dla wszystkich 16 województw
        voivodeships_data = {}
        all_voivodeships = [
            'dolnośląskie', 'kujawsko-pomorskie', 'lubelskie', 'lubuskie',
            'łódzkie', 'małopolskie', 'mazowieckie', 'opolskie',
            'podkarpackie', 'podlaskie', 'pomorskie', 'śląskie',
            'świętokrzyskie', 'warmińsko-mazurskie', 'wielkopolskie',
            'zachodniopomorskie'
        ]
        
        # Inicjalizuj wszystkie województwa z zerami
        for voivodeship in all_voivodeships:
            voivodeships_data[voivodeship] = {
                'production_volume': 0.0,
                'ready_volume': 0.0,
                'total_volume': 0.0,
                'production_value_net': 0.0,
                'ready_value_net': 0.0,
                'total_value_net': 0.0,
                'orders_count': 0
            }
        
        # Mapowanie nazw województw do kluczy (normalizacja)
        voivodeship_mapping = {
            'dolnośląskie': 'dolnośląskie',
            'dolnoslaskie': 'dolnośląskie',
            'kujawsko-pomorskie': 'kujawsko-pomorskie',
            'kujawsko pomorskie': 'kujawsko-pomorskie',
            'lubelskie': 'lubelskie',
            'lubuskie': 'lubuskie', 
            'łódzkie': 'łódzkie',
            'lódzkie': 'łódzkie',
            'lodzkie': 'łódzkie',
            'małopolskie': 'małopolskie',
            'malopolskie': 'małopolskie',
            'mazowieckie': 'mazowieckie',
            'opolskie': 'opolskie',
            'podkarpackie': 'podkarpackie',
            'podlaskie': 'podlaskie',
            'pomorskie': 'pomorskie',
            'śląskie': 'śląskie',
            'slaskie': 'śląskie',
            'sląskie': 'śląskie',
            'świętokrzyskie': 'świętokrzyskie',
            'swietokrzyskie': 'świętokrzyskie',
            'świętokrzyskie': 'świętokrzyskie',
            'warmińsko-mazurskie': 'warmińsko-mazurskie',
            'warminsko-mazurskie': 'warmińsko-mazurskie',
            'warmińsko mazurskie': 'warmińsko-mazurskie',
            'wielkopolskie': 'wielkopolskie',
            'zachodniopomorskie': 'zachodniopomorskie'
        }
        
        # Grupuj dane według województw
        for order in orders:
            delivery_state = order.delivery_state
            if not delivery_state:
                continue
                
            # Normalizuj nazwę województwa
            state_normalized = delivery_state.lower().strip()
            voivodeship_key = voivodeship_mapping.get(state_normalized)
            
            if not voivodeship_key:
                reports_logger.debug("Nieznane województwo", 
                                   delivery_state=delivery_state,
                                   user_email=user_email)
                continue
            
            # Pobierz wartości
            volume = float(order.total_volume or 0)
            value_net = float(order.value_net or 0)
            status_id = order.baselinker_status_id
            
            # Zwiększ licznik zamówień
            voivodeships_data[voivodeship_key]['orders_count'] += 1
            
            # Klasyfikuj według statusu
            if status_id in IN_PRODUCTION_STATUSES:
                voivodeships_data[voivodeship_key]['production_volume'] += volume
                voivodeships_data[voivodeship_key]['production_value_net'] += value_net
            elif status_id in READY_STATUSES:
                voivodeships_data[voivodeship_key]['ready_volume'] += volume
                voivodeships_data[voivodeship_key]['ready_value_net'] += value_net
        
        # Oblicz łączne wartości
        for voivodeship_key in voivodeships_data:
            data = voivodeships_data[voivodeship_key]
            data['total_volume'] = data['production_volume'] + data['ready_volume']
            data['total_value_net'] = data['production_value_net'] + data['ready_value_net']
        
        # Mapowanie kluczy dla frontendu (z polskich do angielskich ID)
        frontend_mapping = {
            'dolnośląskie': 'dolnoslaskie',
            'kujawsko-pomorskie': 'kujawsko-pomorskie', 
            'lubelskie': 'lubelskie',
            'lubuskie': 'lubuskie',
            'łódzkie': 'lodzkie',
            'małopolskie': 'malopolskie',
            'mazowieckie': 'mazowieckie',
            'opolskie': 'opolskie',
            'podkarpackie': 'podkarpackie',
            'podlaskie': 'podlaskie',
            'pomorskie': 'pomorskie',
            'śląskie': 'slaskie',
            'świętokrzyskie': 'swietokrzyskie',
            'warmińsko-mazurskie': 'warminsko-mazurskie',
            'wielkopolskie': 'wielkopolskie',
            'zachodniopomorskie': 'zachodniopomorskie'
        }
        
        # Przekonwertuj klucze dla frontendu
        frontend_data = {}
        for polish_key, data in voivodeships_data.items():
            frontend_key = frontend_mapping.get(polish_key, polish_key)
            frontend_data[frontend_key] = data
        
        # Oblicz statystyki ogólne
        total_production_volume = sum(data['production_volume'] for data in voivodeships_data.values())
        total_ready_volume = sum(data['ready_volume'] for data in voivodeships_data.values())
        total_volume = sum(data['total_volume'] for data in voivodeships_data.values())
        total_orders = sum(data['orders_count'] for data in voivodeships_data.values())
        
        result = {
            'status': 'success',
            'data': frontend_data,
            'summary': {
                'total_production_volume': total_production_volume,
                'total_ready_volume': total_ready_volume,
                'total_volume': total_volume,
                'total_orders': total_orders,
                'processed_orders': len(orders),
                'date_from': date_from.isoformat() if date_from else None,
                'date_to': date_to.isoformat() if date_to else None
            }
        }
        
        reports_logger.info("Zwrócono dane mapy", 
                          summary=result['summary'],
                          user_email=user_email)
        
        return jsonify(result)
        
    except Exception as e:
        reports_logger.error("Błąd API map-statistics", 
                           error=str(e), 
                           user_email=session.get('user_email', 'Nieznany'))
        return jsonify({
            'status': 'error',
            'message': 'Błąd podczas pobierania danych mapy',
            'error': str(e)
        }), 500